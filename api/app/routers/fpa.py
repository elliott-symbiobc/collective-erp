import io
import logging
import os
from datetime import date, datetime, timedelta

import httpx
import psycopg2
import psycopg2.extras
from fastapi import APIRouter, HTTPException, Request, UploadFile, File
from fastapi.responses import Response
from pydantic import BaseModel
from typing import Any

from app.routers.auth import require_admin

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/fpa", tags=["fpa"])


def get_conn():
    conn = psycopg2.connect(os.environ["DATABASE_URL"])
    conn.cursor_factory = psycopg2.extras.RealDictCursor
    return conn


def _fetch_active_model(cur):
    cur.execute("SELECT * FROM fpa_model WHERE is_active = true ORDER BY uploaded_at DESC LIMIT 1")
    return cur.fetchone()


@router.get("/model")
def get_model(request: Request):
    """Return the active financial model. Admin only."""
    require_admin(request)

    conn = get_conn()
    try:
        cur = conn.cursor()
        model = _fetch_active_model(cur)
        cur.execute(
            "SELECT version_number FROM fpa_model_versions WHERE version_number IS NOT NULL ORDER BY created_at DESC LIMIT 1"
        )
        vrow = cur.fetchone()
        cur.close()
    finally:
        conn.close()

    if not model:
        raise HTTPException(status_code=404, detail="No active model found")

    result = dict(model)
    result["version_number"] = vrow["version_number"] if vrow else None
    return result


class ModelUpdate(BaseModel):
    exit_multiple: float | None = None
    exit_year: int | None = None
    start_year: int | None = None
    working_capital: float | None = None
    starting_valuation: float | None = None
    original_equity: float | None = None
    non_cash_equity: float | None = None
    financing_charges_pct: float | None = None
    sba_rate: float | None = None
    bank_debt_rate: float | None = None
    mezz_rate: float | None = None
    hy_rate: float | None = None
    rd_contract_avg_value: float | None = None
    rd_contract_duration_months: int | None = None
    portfolio_contract_avg_value: float | None = None
    portfolio_contract_duration_months: int | None = None
    annual_data: list[dict[str, Any]] | None = None
    monthly_data: list[dict[str, Any]] | None = None
    equity_rounds: list[dict[str, Any]] | None = None
    contract_pipeline: list[dict[str, Any]] | None = None
    headcount_schedule: list[dict[str, Any]] | None = None
    expense_schedule: list[dict[str, Any]] | None = None
    funding_schedule: list[dict[str, Any]] | None = None
    change_summary: str | None = None
    notes: str | None = None
    write_excel: bool = False


@router.patch("/model")
def update_model(body: ModelUpdate, request: Request):
    """Update model assumptions and/or annual data. Admin only."""
    require_admin(request)

    import json

    fields = []
    values = []

    scalar_fields = [
        "exit_multiple", "exit_year", "start_year", "working_capital",
        "starting_valuation", "original_equity", "non_cash_equity",
        "financing_charges_pct", "sba_rate", "bank_debt_rate", "mezz_rate",
        "hy_rate", "rd_contract_avg_value", "rd_contract_duration_months",
        "portfolio_contract_avg_value", "portfolio_contract_duration_months", "notes",
    ]
    for f in scalar_fields:
        val = getattr(body, f)
        if val is not None:
            fields.append(f"{f} = %s")
            values.append(val)

    if body.annual_data is not None:
        fields.append("annual_data = %s")
        values.append(json.dumps(body.annual_data))

    if body.monthly_data is not None:
        fields.append("monthly_data = %s")
        values.append(json.dumps(body.monthly_data))

    if body.equity_rounds is not None:
        fields.append("equity_rounds = %s")
        values.append(json.dumps(body.equity_rounds))

    if body.contract_pipeline is not None:
        fields.append("contract_pipeline = %s")
        values.append(json.dumps(body.contract_pipeline))

    if body.headcount_schedule is not None:
        fields.append("headcount_schedule = %s")
        values.append(json.dumps(body.headcount_schedule))

    if body.expense_schedule is not None:
        fields.append("expense_schedule = %s")
        values.append(json.dumps(body.expense_schedule))

    if body.funding_schedule is not None:
        fields.append("funding_schedule = %s")
        values.append(json.dumps(body.funding_schedule))

    if body.change_summary is not None:
        fields.append("change_summary = %s")
        values.append(body.change_summary)

    if not fields:
        raise HTTPException(status_code=400, detail="No fields to update")

    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute(
            f"UPDATE fpa_model SET {', '.join(fields)} WHERE is_active = true",
            values,
        )
        conn.commit()
        model = _fetch_active_model(cur)
        cur.close()
    finally:
        conn.close()

    if body.write_excel:
        try:
            _write_qbo_actuals_to_excel()
        except Exception:
            logging.exception("Excel write-back after model update failed")

    return dict(model)


@router.get("/scenarios")
def list_scenarios(request: Request):
    """Return saved model scenarios. Admin only."""
    require_admin(request)

    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute(
            "SELECT scenario_id, name, description, created_at, created_by FROM fpa_scenarios ORDER BY created_at DESC"
        )
        rows = cur.fetchall()
        cur.close()
    finally:
        conn.close()

    return [dict(r) for r in rows]


class ScenarioCreate(BaseModel):
    name: str
    description: str | None = None


@router.post("/scenarios", status_code=201)
def create_scenario(body: ScenarioCreate, request: Request):
    """Save current model state as a named scenario. Admin only."""
    user = require_admin(request)

    import json

    conn = get_conn()
    try:
        cur = conn.cursor()
        model = _fetch_active_model(cur)
        if not model:
            raise HTTPException(status_code=404, detail="No active model found")

        snapshot = dict(model)
        # Remove non-serializable UUID and datetime
        snapshot["model_id"] = str(snapshot["model_id"])
        if snapshot.get("uploaded_at"):
            snapshot["uploaded_at"] = snapshot["uploaded_at"].isoformat()

        cur.execute(
            """INSERT INTO fpa_scenarios (name, description, created_by, model_snapshot)
               VALUES (%s, %s, %s, %s) RETURNING scenario_id""",
            (body.name, body.description, user["email"], json.dumps(snapshot)),
        )
        conn.commit()
        scenario_id = cur.fetchone()["scenario_id"]
        cur.close()
    finally:
        conn.close()

    return {"scenario_id": str(scenario_id), "name": body.name}


# ── Excel helpers ─────────────────────────────────────────────────────────────

# Default Office theme accent colours (indices 0-9 of the built-in theme).
_THEME_COLORS = [
    "FFFFFF","000000","E7E6E6","44546A","4472C4",
    "ED7D31","A5A5A5","FFC000","5B9BD5","70AD47",
]

def _resolve_color(color) -> str | None:
    """Openpyxl Color → CSS hex string, or None if transparent/auto."""
    if color is None:
        return None
    t = getattr(color, "type", None)
    if t == "rgb":
        rgb = color.rgb or ""
        if len(rgb) == 8 and rgb[:2] != "00":
            return "#" + rgb[2:]
        if len(rgb) == 6:
            return "#" + rgb
    elif t == "theme":
        idx = getattr(color, "theme", 0) or 0
        tint = getattr(color, "tint", 0.0) or 0.0
        if idx < len(_THEME_COLORS):
            hex_c = _THEME_COLORS[idx]
            if tint == 0:
                return "#" + hex_c
            # Apply tint (lighten/darken)
            r, g, b = int(hex_c[0:2], 16), int(hex_c[2:4], 16), int(hex_c[4:6], 16)
            if tint > 0:
                r = round(r + (255 - r) * tint)
                g = round(g + (255 - g) * tint)
                b = round(b + (255 - b) * tint)
            else:
                r = round(r * (1 + tint))
                g = round(g * (1 + tint))
                b = round(b * (1 + tint))
            return f"#{max(0,min(255,r)):02x}{max(0,min(255,g)):02x}{max(0,min(255,b)):02x}"
    return None


def _fmt_value(cell) -> str:
    """Format a cell value using its number_format."""
    v = cell.value
    if v is None:
        return ""
    fmt = cell.number_format or "General"
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, (int, float)):
        if "%" in fmt:
            return f"{v * 100:.1f}%"
        if "$" in fmt or "£" in fmt or "€" in fmt or "#,##0" in fmt:
            neg = v < 0
            s = f"{abs(v):,.0f}" if ("0" in fmt and ".00" not in fmt) else f"{abs(v):,.2f}"
            return f"(${s})" if neg else f"${s}"
        if v == int(v):
            return f"{int(v):,}"
        return str(round(v, 4))
    import datetime as _dt
    if isinstance(v, (_dt.datetime, _dt.date)):
        return str(v)[:10]
    return str(v)


def _excel_to_html(data: bytes, sheet_name: str | None = None) -> str:
    """Convert one sheet of an Excel workbook to a styled HTML table."""
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter, column_index_from_string

    wb = load_workbook(filename=io.BytesIO(data), data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    # Build merged-cell lookup: {(row,col): (rowspan,colspan) or None if slave}
    merge_info: dict[tuple, tuple | None] = {}
    for merge in ws.merged_cells.ranges:
        rs = merge.max_row - merge.min_row + 1
        cs = merge.max_col - merge.min_col + 1
        for r in range(merge.min_row, merge.max_row + 1):
            for c in range(merge.min_col, merge.max_col + 1):
                if r == merge.min_row and c == merge.min_col:
                    merge_info[(r, c)] = (rs, cs)
                else:
                    merge_info[(r, c)] = None  # slave – skip

    col_widths: dict[int, float] = {}
    for col_letter, cd in ws.column_dimensions.items():
        col_widths[column_index_from_string(col_letter)] = cd.width or 8.43

    rows_html: list[str] = []
    for row in ws.iter_rows():
        rd = ws.row_dimensions.get(row[0].row)
        row_h = f" height=\"{max(1, int((rd.height or 15) * 1.33))}px\"" if rd and rd.height else ""
        cells_html: list[str] = []
        for cell in row:
            key = (cell.row, cell.column)
            mi = merge_info.get(key)
            if mi is None and key in merge_info:
                continue  # slave cell
            style_parts: list[str] = []
            attrs: list[str] = []
            if mi:
                rs, cs = mi
                if rs > 1: attrs.append(f'rowspan="{rs}"')
                if cs > 1: attrs.append(f'colspan="{cs}"')
            # Width
            w = col_widths.get(cell.column, 8.43)
            style_parts.append(f"min-width:{max(20, int(w * 7))}px")
            style_parts.append("max-width:300px")
            # Background
            try:
                bg = _resolve_color(cell.fill.fgColor) if cell.fill and cell.fill.patternType not in (None, "none") else None
            except Exception:
                bg = None
            if bg:
                style_parts.append(f"background:{bg}")
            # Font
            try:
                font = cell.font
                if font:
                    fc = _resolve_color(font.color)
                    if fc: style_parts.append(f"color:{fc}")
                    if font.bold: style_parts.append("font-weight:bold")
                    if font.italic: style_parts.append("font-style:italic")
                    if font.size: style_parts.append(f"font-size:{int(font.size)}px")
            except Exception:
                pass
            # Alignment
            try:
                al = cell.alignment
                if al:
                    hz = al.horizontal
                    if hz in ("right", "center"): style_parts.append(f"text-align:{hz}")
                    elif isinstance(cell.value, (int, float)): style_parts.append("text-align:right")
                elif isinstance(cell.value, (int, float)):
                    style_parts.append("text-align:right")
            except Exception:
                if isinstance(cell.value, (int, float)):
                    style_parts.append("text-align:right")
            # Border bottom (for visual separation)
            try:
                bd = cell.border
                borders = []
                if bd:
                    def _bs(b): return b and b.border_style and b.border_style not in ("none", "thin", None)
                    sides = {"top": bd.top, "right": bd.right, "bottom": bd.bottom, "left": bd.left}
                    for side, b in sides.items():
                        if b and b.border_style:
                            bc = _resolve_color(b.color) or "#ccc"
                            w_px = "2px" if b.border_style in ("medium","thick","double") else "1px"
                            borders.append(f"border-{side}:{w_px} solid {bc}")
                if borders:
                    style_parts.extend(borders)
            except Exception:
                pass

            style = ";".join(style_parts)
            attr_str = (" " + " ".join(attrs)) if attrs else ""
            val = _fmt_value(cell)
            import html as _html
            cells_html.append(f'<td{attr_str} style="{style}" title="{_html.escape(val)}">{_html.escape(val)}</td>')
        rows_html.append(f"<tr{row_h}>{''.join(cells_html)}</tr>")

    sheet_names_js = ", ".join(f'"{s}"' for s in wb.sheetnames)
    return f"""<!doctype html><html><head><meta charset="utf-8">
<style>
  body{{margin:0;font-family:Calibri,Arial,sans-serif;font-size:12px;background:#fff}}
  table{{border-collapse:collapse;white-space:nowrap}}
  td{{padding:2px 6px;overflow:hidden;text-overflow:ellipsis;border:1px solid #e5e7eb;vertical-align:middle}}
</style></head><body>
<table>{''.join(rows_html)}</table>
</body></html>"""


# Label → model field mapping for Excel parsing
_METRIC_LABELS: dict[str, str] = {
    "revenue": "revenue", "total revenue": "revenue", "net revenue": "revenue",
    "gross revenue": "revenue", "sales": "revenue",
    "ebitda": "ebitda", "operating income": "ebitda",
    "total expenses": "total_opex", "total opex": "total_opex",
    "operating expenses": "total_opex", "total operating expenses": "total_opex",
    "payroll": "exec_comp", "total payroll": "exec_comp", "wages": "exec_comp",
    "salaries": "exec_comp", "compensation": "exec_comp",
    "software": "software", "software & apps": "software",
    "office expense": "office_expense", "office expenses": "office_expense",
    "facilities": "office_expense", "rent": "office_expense",
    "legal": "legal_accounting", "legal & accounting": "legal_accounting",
    "gross margin": "gross_margin", "gross profit": "gross_margin",
}


def _parse_excel_cell_map(data: bytes) -> dict:
    """
    Scan uploaded Excel for an annual P&L structure.
    Returns {sheet, year_cols, metrics, annual_data}.
    """
    from openpyxl import load_workbook
    wb = load_workbook(filename=io.BytesIO(data), data_only=True)

    best: dict | None = None
    best_score = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Scan first 20 rows for year headers (integers 2020-2040)
        year_cols: dict[int, int] = {}  # year → 1-based col index
        header_row = None
        for ri, row in enumerate(ws.iter_rows(max_row=30, values_only=True), 1):
            years_in_row = {
                ci: v for ci, v in enumerate(row, 1)
                if isinstance(v, (int, float)) and 2020 <= v <= 2040
            }
            if len(years_in_row) >= 2:
                year_cols = {int(v): ci for ci, v in years_in_row.items()}
                header_row = ri
                break

        if not year_cols:
            continue

        # Scan subsequent rows for metric labels
        metrics: dict[str, dict] = {}  # field → {row, col_label, year_values}
        for ri, row in enumerate(ws.iter_rows(min_row=header_row + 1, max_row=header_row + 80, values_only=True), header_row + 1):
            for ci, val in enumerate(row, 1):
                if not isinstance(val, str):
                    continue
                label_lower = val.strip().lower()
                if label_lower in _METRIC_LABELS:
                    field = _METRIC_LABELS[label_lower]
                    if field not in metrics:
                        metrics[field] = {"row": ri, "label_col": ci, "year_values": {}}
                        for yr, yr_col in year_cols.items():
                            cell_val = ws.cell(ri, yr_col).value
                            if isinstance(cell_val, (int, float)):
                                metrics[field]["year_values"][yr] = cell_val

        score = len(metrics) * len(year_cols)
        if score > best_score:
            best_score = score
            best = {"sheet": sheet_name, "year_cols": year_cols, "metrics": metrics}

    if not best or not best.get("metrics"):
        return {}

    # Build annual_data list from parsed metrics
    years = sorted(best["year_cols"].keys())
    annual_data = []
    for yr in years:
        row_d: dict[str, Any] = {"year": yr}
        for field, info in best["metrics"].items():
            v = info["year_values"].get(yr)
            if v is not None:
                row_d[field] = round(float(v), 2)
        # Derive ebitda if missing
        if "ebitda" not in row_d and "revenue" in row_d and "total_opex" in row_d:
            row_d["ebitda"] = round(row_d["revenue"] - row_d["total_opex"], 2)
        if row_d.keys() - {"year"}:
            annual_data.append(row_d)

    best["annual_data"] = annual_data
    return best


def _apply_actuals_to_excel(data: bytes, cell_map: dict, actuals_by_field: dict[str, dict[int, float]]) -> bytes:
    """
    Write actual values into the Excel at the mapped cells.
    Only overwrites non-formula cells. Returns updated workbook bytes.
    actuals_by_field: {field: {year: value}}
    """
    from openpyxl import load_workbook
    wb = load_workbook(filename=io.BytesIO(data))
    sheet_name = cell_map.get("sheet")
    if not sheet_name or sheet_name not in wb.sheetnames:
        return data
    ws = wb[sheet_name]
    year_cols: dict[int, int] = {int(k): v for k, v in cell_map.get("year_cols", {}).items()}
    metrics: dict[str, dict] = cell_map.get("metrics", {})

    for field, year_vals in actuals_by_field.items():
        if field not in metrics:
            continue
        row_idx = metrics[field]["row"]
        for year, value in year_vals.items():
            col_idx = year_cols.get(int(year))
            if not col_idx:
                continue
            cell = ws.cell(row_idx, col_idx)
            # Skip formula cells
            if isinstance(cell.value, str) and cell.value.startswith("="):
                continue
            cell.value = round(value, 2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _parse_op_model_summary(data: bytes) -> dict | None:
    """
    Parse the 'Op Model Summary' sheet specifically.
    Structure: each year = 17 cols (12 months + Q1-Q4 + Annual).
    Returns {annual_data, monthly_data} or None if sheet not found.
    """
    from openpyxl import load_workbook
    wb = load_workbook(filename=io.BytesIO(data), data_only=True)
    if "Op Model Summary" not in wb.sheetnames:
        return None
    ws = wb["Op Model Summary"]

    YEAR_START_COL = 7   # first month column (1-based)
    BLOCK = 17            # columns per year (12 months + Q1-Q4 + Annual)
    ANNUAL_OFFSET = 16    # 0-based offset within block for annual total

    MONTH_NAMES = ["January","February","March","April","May","June",
                   "July","August","September","October","November","December"]

    # Row definitions (non-cumulative section)
    ROW_MAP = {
        "rd_contract_revenue": 16,
        "portfolio_revenue": 17,
        "maintenance_revenue": 18,
        "grant_revenue": 19,
        "total_revenue": 20,
        "direct_expenses": 24,
        "gross_margin": 26,
        "tech_comp": 31,
        "exec_comp": 32,
        "ga_comp": 33,
        "sales_comp": 34,
        "office_expense": 35,
        "legal_accounting": 36,
        "software": 37,
        "total_opex": 38,
        "ebitda": 41,
    }

    def _v(row_n: int, col: int) -> float:
        v = ws.cell(row_n, col).value
        return round(float(v), 2) if isinstance(v, (int, float)) else 0.0

    annual_data = []
    monthly_data = []

    for i, year in enumerate(range(2025, 2035)):
        block_start = YEAR_START_COL + i * BLOCK
        annual_col = block_start + ANNUAL_OFFSET

        # Annual totals
        entry: dict[str, Any] = {"year": year}
        for field, row_n in ROW_MAP.items():
            entry[field] = _v(row_n, annual_col)
        # Alias for backward compat
        entry["revenue"] = entry["total_revenue"]
        annual_data.append(entry)

        # Monthly data (first 12 cols of the block)
        for m in range(12):
            col = block_start + m
            mo: dict[str, Any] = {"year": year, "month": m + 1, "month_name": MONTH_NAMES[m]}
            for field, row_n in ROW_MAP.items():
                mo[field] = _v(row_n, col)
            monthly_data.append(mo)

    return {"annual_data": annual_data, "monthly_data": monthly_data}


@router.post("/model/upload-excel")
async def upload_excel(request: Request, file: UploadFile = File(...)):
    """Store an uploaded Excel model file, parse model data, update fpa_model. Admin only."""
    require_admin(request)
    if not file.filename.endswith((".xlsx", ".xlsm", ".xls")):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx/.xlsm/.xls) are accepted")
    data = await file.read()
    if len(data) > 20 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large (max 20 MB)")

    # Try specialized Op Model Summary parser first, fall back to generic
    parsed = _parse_op_model_summary(data)
    if parsed:
        annual_data = parsed["annual_data"]
        monthly_data = parsed["monthly_data"]
        cell_map = {}
    else:
        cell_map = _parse_excel_cell_map(data)
        annual_data = cell_map.pop("annual_data", None)
        monthly_data = []

    import json as _json
    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM fpa_uploaded_excel")
        cur.execute(
            "INSERT INTO fpa_uploaded_excel (filename, data, cell_map) VALUES (%s, %s, %s)",
            (file.filename, psycopg2.Binary(data), _json.dumps(cell_map) if cell_map else None),
        )
        if annual_data:
            cur.execute(
                "UPDATE fpa_model SET annual_data = %s, monthly_data = %s WHERE is_active = true",
                (_json.dumps(annual_data), _json.dumps(monthly_data)),
            )
        conn.commit()
        cur.close()
    finally:
        conn.close()

    return {
        "status": "uploaded",
        "filename": file.filename,
        "size": len(data),
        "parsed_years": [d["year"] for d in (annual_data or [])],
        "parsed_metrics": list(ROW_MAP.keys() if parsed else cell_map.get("metrics", {}).keys()),
    }


@router.get("/model/monthly")
def get_model_monthly(request: Request):
    """Return monthly model data parsed from the uploaded Excel. Admin only."""
    require_admin(request)
    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute("SELECT monthly_data FROM fpa_model WHERE is_active = true LIMIT 1")
        row = cur.fetchone()
        cur.close()
    finally:
        conn.close()
    if not row or not row["monthly_data"]:
        return []
    return row["monthly_data"]


@router.get("/model/excel-info")
def excel_info(request: Request):
    """Return metadata about the uploaded Excel file. Admin only."""
    require_admin(request)
    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute("SELECT filename, uploaded_at, octet_length(data) AS size FROM fpa_uploaded_excel ORDER BY id DESC LIMIT 1")
        row = cur.fetchone()
        cur.close()
    finally:
        conn.close()
    if not row:
        return {"uploaded": False}
    return {"uploaded": True, "filename": row["filename"], "uploaded_at": row["uploaded_at"].isoformat(), "size": row["size"]}


@router.get("/model/excel-raw")
def excel_raw(request: Request):
    """Serve the uploaded Excel file for download. Admin only."""
    require_admin(request)
    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute("SELECT filename, data FROM fpa_uploaded_excel ORDER BY id DESC LIMIT 1")
        row = cur.fetchone()
        cur.close()
    finally:
        conn.close()
    if not row:
        raise HTTPException(status_code=404, detail="No uploaded Excel file found")
    return Response(
        content=bytes(row["data"]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'inline; filename="{row["filename"]}"'},
    )


@router.get("/model/excel-html")
def excel_html_view(request: Request, sheet: str | None = None):
    """Return styled HTML for one sheet of the uploaded Excel. Admin only."""
    require_admin(request)
    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute("SELECT data, filename FROM fpa_uploaded_excel ORDER BY id DESC LIMIT 1")
        row = cur.fetchone()
        cur.close()
    finally:
        conn.close()
    if not row:
        raise HTTPException(status_code=404, detail="No uploaded Excel file found")
    html = _excel_to_html(bytes(row["data"]), sheet)
    from fastapi.responses import HTMLResponse
    return HTMLResponse(content=html)


@router.get("/model/excel-sheets")
def excel_sheets(request: Request):
    """Return list of sheet names in the uploaded Excel. Admin only."""
    require_admin(request)
    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute("SELECT data FROM fpa_uploaded_excel ORDER BY id DESC LIMIT 1")
        row = cur.fetchone()
        cur.close()
    finally:
        conn.close()
    if not row:
        return []
    from openpyxl import load_workbook
    wb = load_workbook(filename=io.BytesIO(bytes(row["data"])), read_only=True)
    return wb.sheetnames


@router.get("/model/export")
def export_model(request: Request):
    """Return the uploaded Excel if available, otherwise generate from model. Admin only."""
    require_admin(request)

    # Prefer the uploaded file
    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute("SELECT filename, data FROM fpa_uploaded_excel ORDER BY id DESC LIMIT 1")
        row = cur.fetchone()
        cur.close()
    finally:
        conn.close()
    if row:
        return Response(
            content=bytes(row["data"]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{row["filename"]}"'},
        )

    # Fall back to generated model
    require_admin(request)

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    conn = get_conn()
    try:
        cur = conn.cursor()
        model = _fetch_active_model(cur)
        cur.close()
    finally:
        conn.close()

    if not model:
        raise HTTPException(status_code=404, detail="No active model found")

    annual_data = model["annual_data"] or []

    wb = openpyxl.Workbook()

    NAVY = "1F3864"
    BLUE_INPUT = "0000FF"
    BLACK = "000000"
    GREEN = "006400"
    YELLOW = "FFFF00"
    LIGHT_GRAY = "F2F2F2"
    WHITE = "FFFFFF"

    def navy_header(cell, text):
        cell.value = text
        cell.font = Font(bold=True, color=WHITE, size=11)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center")

    def blue_input(cell, value, fmt=None):
        cell.value = value
        cell.font = Font(color=BLUE_INPUT)
        cell.fill = PatternFill("solid", fgColor=YELLOW)
        if fmt:
            cell.number_format = fmt

    def bold_label(cell, text):
        cell.value = text
        cell.font = Font(bold=True, size=11)

    # ── SHEET 1: Returns Model ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Returns Model"
    ws1.column_dimensions["A"].width = 35
    ws1.column_dimensions["B"].width = 20

    ws1["A1"].value = "Collective ERP — Returns Model"
    ws1["A1"].font = Font(bold=True, size=14)

    bold_label(ws1["A3"], "EXIT ASSUMPTIONS")

    assumptions = [
        ("Exit Multiple of EBITDA", model["exit_multiple"], '0.0"x"'),
        ("Exit Year (from start)", model["exit_year"], "0"),
        ("Starting Valuation", model["starting_valuation"], '$#,##0'),
        ("Original Equity Investment", model["original_equity"], '$#,##0'),
        ("Non-Cash Equity Value", model["non_cash_equity"], '$#,##0'),
        ("Working Capital", model["working_capital"], '$#,##0'),
    ]
    for i, (label, val, fmt) in enumerate(assumptions, start=4):
        ws1.cell(i, 1).value = label
        blue_input(ws1.cell(i, 2), val, fmt)

    bold_label(ws1["A11"], "IMPLIED RETURNS")

    exit_ebitda = 0
    for yr in annual_data:
        if isinstance(yr, dict) and yr.get("year") == model["start_year"] + model["exit_year"] - 1:
            exit_ebitda = yr.get("ebitda", 0)
    # fallback: last year
    if exit_ebitda == 0 and annual_data:
        exit_ebitda = annual_data[-1].get("ebitda", 0)

    ws1["A12"].value = f"Exit Year EBITDA ({model['start_year'] + model['exit_year'] - 1})"
    ws1["B12"].value = exit_ebitda
    ws1["B12"].number_format = '$#,##0'

    ws1["A13"].value = "Implied Exit Valuation"
    ws1["B13"].value = f"=B4*B12"
    ws1["B13"].number_format = '$#,##0'
    ws1["B13"].font = Font(bold=True, color=GREEN)

    ws1["A14"].value = "Implied MOIC"
    ws1["B14"].value = "=B13/B7"
    ws1["B14"].number_format = '0.0"x"'
    ws1["B14"].font = Font(bold=True, color=GREEN)

    bold_label(ws1["A16"], "DEBT FINANCING RATES")
    rates = [
        ("SBA Loan Rate", model["sba_rate"]),
        ("Bank Debt Rate", model["bank_debt_rate"]),
        ("Mezzanine Rate", model["mezz_rate"]),
        ("High Yield Rate", model["hy_rate"]),
    ]
    for i, (label, rate) in enumerate(rates, start=17):
        ws1.cell(i, 1).value = label
        blue_input(ws1.cell(i, 2), rate, "0.0%")

    bold_label(ws1["A22"], "EQUITY ROUNDS")
    for col, hdr in [(1, "Round"), (2, "Amount"), (3, "Date"), (4, "Cumulative"), (5, "Notes")]:
        c = ws1.cell(22, col)
        navy_header(c, hdr)
        ws1.column_dimensions[get_column_letter(col)].width = 14 if col > 1 else 10

    cumulative = 0
    for i, rnd in enumerate(model["equity_rounds"] or [], start=23):
        cumulative += rnd.get("amount", 0)
        ws1.cell(i, 1).value = rnd.get("round", i - 22)
        ws1.cell(i, 2).value = rnd.get("amount", 0)
        ws1.cell(i, 2).number_format = '$#,##0'
        ws1.cell(i, 3).value = rnd.get("date", "")
        ws1.cell(i, 4).value = cumulative
        ws1.cell(i, 4).number_format = '$#,##0'
        ws1.cell(i, 5).value = rnd.get("notes", "")

    # ── SHEET 2: P&L Summary ────────────────────────────────────────────────
    ws2 = wb.create_sheet("P&L Summary")
    ws2.column_dimensions["A"].width = 32

    ws2["A1"].value = "Collective ERP — P&L Summary ($)"
    ws2["A1"].font = Font(bold=True, size=13)

    years = [d["year"] for d in annual_data if isinstance(d, dict)]
    navy_header(ws2["A3"], "Metric")
    for j, yr in enumerate(years, start=2):
        col = get_column_letter(j)
        ws2.column_dimensions[col].width = 14
        navy_header(ws2.cell(3, j), yr)

    MONEY = '$#,##0_);($#,##0)'
    PCT = '0.0%;(0.0%)'

    rows_def = [
        ("Revenue", "revenue", True, MONEY),
        ("Gross Margin", "gross_margin", False, MONEY),
        ("Gross Margin %", "_gm_pct", False, PCT),
        (None, None, None, None),
        ("  Technical Compensation", "tech_comp", True, MONEY),
        ("  Executive Compensation", "exec_comp", True, MONEY),
        ("  G&A Compensation", "ga_comp", True, MONEY),
        ("  Sales Compensation", "sales_comp", True, MONEY),
        ("  Office Expense", "office_expense", True, MONEY),
        ("  Legal & Accounting", "legal_accounting", True, MONEY),
        ("  Software", "software", True, MONEY),
        ("Total OpEx", "total_opex", False, MONEY),
        (None, None, None, None),
        ("EBITDA", "ebitda", False, MONEY),
        ("EBITDA Margin %", "_ebitda_pct", False, PCT),
    ]

    for row_offset, (label, field, is_input, fmt) in enumerate(rows_def, start=4):
        if label is None:
            continue
        r = row_offset
        label_cell = ws2.cell(r, 1)
        label_cell.value = label
        is_bold_row = label in ("Total OpEx", "EBITDA", "Gross Margin", "Revenue")
        if is_bold_row:
            label_cell.font = Font(bold=True)
            label_cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)

        for j, yr_data in enumerate(annual_data, start=2):
            if not isinstance(yr_data, dict):
                continue
            c = ws2.cell(r, j)
            rev = yr_data.get("revenue", 1) or 1

            if field == "_gm_pct":
                c.value = yr_data.get("gross_margin", 0) / rev
                c.number_format = PCT
                c.font = Font(color=BLACK)
            elif field == "_ebitda_pct":
                ebitda = yr_data.get("ebitda", 0)
                c.value = ebitda / rev if rev else 0
                c.number_format = PCT
                c.font = Font(bold=True, color=BLACK)
                c.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
            elif field and field in yr_data:
                val = yr_data[field]
                # expense rows stored as positive, display as negative
                if is_input and field not in ("revenue", "gross_margin"):
                    val = -abs(val)
                c.value = val
                c.number_format = fmt
                c.font = Font(
                    color=BLUE_INPUT if is_input else BLACK,
                    bold=is_bold_row,
                )
                if is_bold_row:
                    c.fill = PatternFill("solid", fgColor=LIGHT_GRAY)

    # ── SHEET 3: Revenue Assumptions ────────────────────────────────────────
    ws3 = wb.create_sheet("Revenue Assumptions")
    ws3.column_dimensions["A"].width = 35
    ws3.column_dimensions["B"].width = 20

    ws3["A1"].value = "Revenue Model Assumptions"
    ws3["A1"].font = Font(bold=True, size=13)

    bold_label(ws3["A3"], "R&D CONTRACTS")
    ws3["A4"].value = "Average Contract Value"
    blue_input(ws3["B4"], model["rd_contract_avg_value"], '$#,##0')
    ws3["A5"].value = "Duration (months)"
    blue_input(ws3["B5"], model["rd_contract_duration_months"])
    ws3["A6"].value = "Revenue per Contract"
    ws3["B6"].value = "=B4*B5"
    ws3["B6"].number_format = '$#,##0'

    bold_label(ws3["A8"], "PORTFOLIO CONTRACTS")
    ws3["A9"].value = "Average Contract Value"
    blue_input(ws3["B9"], model["portfolio_contract_avg_value"], '$#,##0')
    ws3["A10"].value = "Duration (months)"
    blue_input(ws3["B10"], model["portfolio_contract_duration_months"])
    ws3["A11"].value = "Revenue per Contract"
    ws3["B11"].value = "=B9*B10"
    ws3["B11"].number_format = '$#,##0'

    bold_label(ws3["A13"], "ANNUAL CONTRACT PROJECTIONS")
    for col, hdr in [(1, "Year"), (2, "New R&D"), (3, "New Portfolio"), (4, "Total")]:
        navy_header(ws3.cell(14, col), hdr)
        ws3.column_dimensions[get_column_letter(col)].width = 16

    new_contracts = [
        (2025, 4, 1), (2026, 4, 12), (2027, 4, 42),
        (2028, 4, 150), (2029, 4, 234), (2030, 4, 282),
        (2031, 4, 330), (2032, 4, 378), (2033, 4, 426),
        (2034, 4, 474),
    ]
    for i, (yr, rd, port) in enumerate(new_contracts, start=15):
        ws3.cell(i, 1).value = yr
        ws3.cell(i, 2).value = rd
        ws3.cell(i, 2).font = Font(color=BLUE_INPUT)
        ws3.cell(i, 3).value = port
        ws3.cell(i, 3).font = Font(color=BLUE_INPUT)
        ws3.cell(i, 4).value = rd + port

    # ── SHEET 4: Balance Sheet ───────────────────────────────────────────────
    ws4 = wb.create_sheet("Balance Sheet")
    ws4.column_dimensions["A"].width = 8
    ws4.column_dimensions["B"].width = 35
    ws4.column_dimensions["C"].width = 18

    ws4["B1"].value = "Collective ERP LLC"
    ws4["B1"].font = Font(bold=True, size=13)
    ws4["B2"].value = "Pro Forma Balance Sheet"
    ws4["B3"].value = "As of March 15, 2025"

    bold_label(ws4["B5"], "ASSETS")
    bold_label(ws4["B6"], "Current Assets")
    ws4["B7"].value = "Bank Accounts"
    ws4["C7"].value = 120000
    ws4["C7"].number_format = '$#,##0'

    bold_label(ws4["B9"], "Fixed Assets")
    assets = [
        ("Lab furniture", 17150),
        ("Refrigeration/cooling", 1700),
        ("Instruments & installation", 36034),
        ("Lab equipment for applications", 500),
        ("Glassware and supplies", 779),
    ]
    for i, (name, val) in enumerate(assets, start=10):
        ws4.cell(i, 2).value = name
        ws4.cell(i, 3).value = val
        ws4.cell(i, 3).number_format = '$#,##0'

    total_row = 10 + len(assets)
    ws4.cell(total_row, 2).value = "Total Fixed Assets"
    ws4.cell(total_row, 2).font = Font(bold=True)
    ws4.cell(total_row, 3).value = f"=SUM(C10:C{total_row-1})"
    ws4.cell(total_row, 3).number_format = '$#,##0'
    ws4.cell(total_row, 3).font = Font(bold=True)

    ws4.cell(total_row + 1, 2).value = "Total Assets"
    ws4.cell(total_row + 1, 2).font = Font(bold=True, size=11)
    ws4.cell(total_row + 1, 3).value = f"=C7+C{total_row}"
    ws4.cell(total_row + 1, 3).number_format = '$#,##0'
    ws4.cell(total_row + 1, 3).font = Font(bold=True, size=11)

    eq_row = total_row + 3
    bold_label(ws4.cell(eq_row, 2), "EQUITY & LIABILITIES")
    ws4.cell(eq_row + 1, 2).value = "Equity Investment (Round 1)"
    ws4.cell(eq_row + 1, 3).value = 120000
    ws4.cell(eq_row + 1, 3).number_format = '$#,##0'
    ws4.cell(eq_row + 2, 2).value = "Total Liabilities"
    ws4.cell(eq_row + 2, 3).value = 0
    ws4.cell(eq_row + 2, 3).number_format = '$#,##0'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Symbio_Financial_Model.xlsx"},
    )


# ── Calculation Engine ────────────────────────────────────────────────────────

MONTH_NAMES = [
    "January","February","March","April","May","June",
    "July","August","September","October","November","December",
]

def _run_engine(model: dict) -> dict:
    """
    Compute monthly_data, annual_data, and audit_data from driver sheets.
    Falls back to existing monthly_data for non-driver cost lines (office, legal, software).
    """
    from datetime import date as _date

    contracts = model.get("contract_pipeline") or []
    headcount = model.get("headcount_schedule") or []
    expenses  = model.get("expense_schedule") or []
    funding   = model.get("funding_schedule") or []
    import datetime as _datetime_mod
    start_year = int(model.get("start_year") or _datetime_mod.date.today().year)

    # Pre-index non-dilutive lump-sum funding by (year, month) for O(1) lookup
    _nondilutive_types = {"grant", "sbir", "loan"}
    lump_fund: dict[tuple, list] = {}
    for f in funding:
        if f.get("disbursement", "lump_sum") != "lump_sum":
            continue
        if f.get("type", "equity") not in _nondilutive_types:
            continue
        try:
            fd = _dt.strptime(f.get("date", "")[:10], "%Y-%m-%d").date()
            lump_fund.setdefault((fd.year, fd.month), []).append(f)
        except Exception:
            pass

    # Build lookup from existing monthly_data as fallback when no expense_schedule entries
    existing_monthly: dict[tuple, dict] = {}
    for m in (model.get("monthly_data") or []):
        existing_monthly[(int(m.get("year", 0)), int(m.get("month", 0)))] = m

    monthly_data = []
    audit_data = []

    for year in range(start_year, start_year + 10):
        for month in range(1, 13):
            current_date = _date(year, month, 1)
            current_ym = year * 12 + month - 1

            # ── Revenue from contracts ────────────────────────────────────────
            rd_rev = portfolio_rev = maintenance_rev = grant_rev = 0.0
            consumable_rev = licensing_rev = 0.0
            rev_components = []

            _recurring_types = {"consumable_recurring", "licensing_recurring"}

            for c in contracts:
                if c.get("status") not in ("active", "pipeline"):
                    continue
                try:
                    from datetime import datetime as _dt
                    sd = _dt.strptime(c.get("start_date", "2025-01-01")[:10], "%Y-%m-%d")
                except Exception:
                    continue
                ctype = c.get("type", "rd_contract")
                probability = float(c.get("probability") or 1.0)

                if ctype in _recurring_types:
                    # Recurring revenue: monthly_amount * (1 + annual_increase)^years * prob
                    if current_date < _date(sd.year, sd.month, 1):
                        continue
                    end_date_str = c.get("end_date", "")
                    if end_date_str:
                        try:
                            ed = _dt.strptime(end_date_str[:10], "%Y-%m-%d").date()
                            if current_date > ed:
                                continue
                        except Exception:
                            pass
                    monthly_base = float(c.get("monthly_amount") or 0)
                    annual_increase = float(c.get("annual_increase_pct") or 0)
                    years_since_deploy = max(0, year - sd.year)
                    effective_monthly = monthly_base * ((1 + annual_increase) ** years_since_deploy) * probability
                    if ctype == "consumable_recurring":
                        consumable_rev += effective_monthly
                    else:
                        licensing_rev += effective_monthly
                    inc_note = f" × {1 + annual_increase:.4g}^{years_since_deploy}yr" if years_since_deploy > 0 and annual_increase else ""
                    rev_components.append({
                        "contract_id": c.get("id", ""),
                        "name": c.get("name", ""),
                        "type": ctype,
                        "monthly_value": round(effective_monthly, 2),
                        "formula": f"${monthly_base:,.0f}/mo{inc_note} × {probability:.0%}",
                    })
                else:
                    duration = max(1, int(c.get("duration_months") or 1))
                    total_value = float(c.get("total_value") or 0)
                    start_ym = sd.year * 12 + sd.month - 1
                    end_ym = start_ym + duration - 1

                    if start_ym <= current_ym <= end_ym:
                        contrib = total_value / duration * probability
                        if ctype == "rd_contract":
                            rd_rev += contrib
                        elif ctype == "portfolio":
                            portfolio_rev += contrib
                        elif ctype == "maintenance":
                            maintenance_rev += contrib
                        elif ctype == "grant":
                            grant_rev += contrib
                        rev_components.append({
                            "contract_id": c.get("id", ""),
                            "name": c.get("name", ""),
                            "type": ctype,
                            "monthly_value": round(contrib, 2),
                            "formula": f"${total_value:,.0f} ÷ {duration}mo × {probability:.0%}",
                        })

                    # post-contract recurring revenue (supports independent consumable + licensing amounts)
                    if c.get("recurring_after_fixed"):
                        rec_years = max(0, int(c.get("recurring_years") or 0))
                        if rec_years > 0:
                            rec_start_ym = end_ym + 1
                            rec_end_ym = rec_start_ym + rec_years * 12 - 1
                            if rec_start_ym <= current_ym <= rec_end_ym:
                                # consumable portion (migrate from old single-type field)
                                rec_consumable = float(c.get("recurring_consumable_amount") or
                                    (c.get("recurring_monthly_amount") if c.get("recurring_type") != "licensing_recurring" else 0) or 0)
                                # licensing portion
                                rec_licensing = float(c.get("recurring_licensing_amount") or
                                    (c.get("recurring_monthly_amount") if c.get("recurring_type") == "licensing_recurring" else 0) or 0)
                                if rec_consumable > 0:
                                    consumable_rev += rec_consumable * probability
                                    rev_components.append({
                                        "contract_id": c.get("id", ""),
                                        "name": f"{c.get('name', '')} (post-contract consumables)",
                                        "type": "consumable_recurring",
                                        "monthly_value": round(rec_consumable * probability, 2),
                                        "formula": f"${rec_consumable:,.0f}/mo × {probability:.0%} ({rec_years}yr post-contract)",
                                    })
                                if rec_licensing > 0:
                                    licensing_rev += rec_licensing * probability
                                    rev_components.append({
                                        "contract_id": c.get("id", ""),
                                        "name": f"{c.get('name', '')} (post-contract licensing)",
                                        "type": "licensing_recurring",
                                        "monthly_value": round(rec_licensing * probability, 2),
                                        "formula": f"${rec_licensing:,.0f}/mo × {probability:.0%} ({rec_years}yr post-contract)",
                                    })

            # ── Non-dilutive funding → grant_revenue ─────────────────────────
            fund_components = []
            # Monthly-disbursed grants/SBIR
            for f in funding:
                ftype = f.get("type", "equity")
                if ftype not in _nondilutive_types:
                    continue
                if f.get("disbursement", "lump_sum") != "monthly":
                    continue
                try:
                    fs = _dt.strptime(f.get("start_date", "")[:10], "%Y-%m-%d").date()
                except Exception:
                    continue
                fe = None
                if f.get("end_date"):
                    try:
                        fe = _dt.strptime(f["end_date"][:10], "%Y-%m-%d").date()
                    except Exception:
                        pass
                if current_date < fs:
                    continue
                if fe and current_date > fe:
                    continue
                monthly_amt = float(f.get("monthly_amount") or 0)
                if ftype == "loan":
                    pass  # loans don't hit P&L revenue
                else:
                    grant_rev += monthly_amt
                fund_components.append({
                    "id": f.get("id", ""),
                    "name": f.get("name", ""),
                    "type": ftype,
                    "monthly_value": round(monthly_amt, 2),
                    "formula": f"${monthly_amt:,.0f}/mo",
                    "hits_pnl": ftype != "loan",
                })
            # Lump-sum grants/SBIR in this specific month
            for f in lump_fund.get((year, month), []):
                ftype = f.get("type", "grant")
                lump = float(f.get("amount") or 0)
                if ftype != "loan":
                    grant_rev += lump
                fund_components.append({
                    "id": f.get("id", ""),
                    "name": f.get("name", ""),
                    "type": ftype,
                    "monthly_value": round(lump, 2),
                    "formula": f"${lump:,.0f} lump sum",
                    "hits_pnl": ftype != "loan",
                })

            total_revenue = rd_rev + portfolio_rev + maintenance_rev + grant_rev + consumable_rev + licensing_rev

            # ── Expenses from headcount ───────────────────────────────────────
            tech_comp = exec_comp = ga_comp = sales_comp = 0.0
            hc_components = []

            for emp in headcount:
                try:
                    from datetime import datetime as _dt
                    emp_start = _dt.strptime(emp.get("start_date", "2025-01-01")[:10], "%Y-%m-%d").date()
                except Exception:
                    continue
                emp_end = None
                if emp.get("end_date"):
                    try:
                        emp_end = _dt.strptime(emp["end_date"][:10], "%Y-%m-%d").date()
                    except Exception:
                        pass
                if current_date < emp_start:
                    continue
                if emp_end and current_date > emp_end:
                    continue
                annual_salary = float(emp.get("annual_salary") or 0)
                benefits_pct = float(emp.get("benefits_pct") or 0.25)
                annual_raise_pct = float(emp.get("annual_raise_pct") or 0)
                years_since_start = max(0, year - emp_start.year)
                salary_overrides = emp.get("salary_overrides") or {}
                if str(year) in salary_overrides:
                    effective_salary = float(salary_overrides[str(year)])
                else:
                    effective_salary = annual_salary * ((1 + annual_raise_pct) ** years_since_start)
                monthly_cost = effective_salary / 12 * (1 + benefits_pct)
                dept = emp.get("department", "ga")
                if dept == "tech":
                    tech_comp += monthly_cost
                elif dept == "exec":
                    exec_comp += monthly_cost
                elif dept == "ga":
                    ga_comp += monthly_cost
                elif dept == "sales":
                    sales_comp += monthly_cost
                raise_note = f" × {1 + annual_raise_pct:.4g}^{years_since_start}yr" if years_since_start > 0 and annual_raise_pct else ""
                hc_components.append({
                    "employee_id": emp.get("id", ""),
                    "role": emp.get("role", ""),
                    "department": dept,
                    "monthly_cost": round(monthly_cost, 2),
                    "formula": f"${annual_salary:,.0f}{raise_note}/12 × (1 + {benefits_pct:.0%} benefits)",
                })

            # ── Expense schedule drivers ──────────────────────────────────────
            office_expense = legal_accounting = software = direct_expenses = 0.0
            exp_components = []
            if expenses:
                for exp in expenses:
                    try:
                        exp_start = _dt.strptime(exp.get("start_date", "2025-01-01")[:10], "%Y-%m-%d").date()
                    except Exception:
                        continue
                    exp_end = None
                    if exp.get("end_date"):
                        try:
                            exp_end = _dt.strptime(exp["end_date"][:10], "%Y-%m-%d").date()
                        except Exception:
                            pass
                    if current_date < exp_start:
                        continue
                    if exp_end and current_date > exp_end:
                        continue
                    monthly_base = float(exp.get("monthly_amount") or 0)
                    annual_increase = float(exp.get("annual_increase_pct") or 0)
                    years_since_start = max(0, year - exp_start.year)
                    amount_overrides = exp.get("amount_overrides") or {}
                    if str(year) in amount_overrides:
                        effective_monthly = float(amount_overrides[str(year)])
                    else:
                        effective_monthly = monthly_base * ((1 + annual_increase) ** years_since_start)
                    cat = exp.get("category", "office_expense")
                    if cat == "office_expense":
                        office_expense += effective_monthly
                    elif cat == "legal_accounting":
                        legal_accounting += effective_monthly
                    elif cat == "software":
                        software += effective_monthly
                    elif cat == "direct_expenses":
                        direct_expenses += effective_monthly
                    increase_note = f" × {1 + annual_increase:.4g}^{years_since_start}yr" if years_since_start > 0 and annual_increase else ""
                    exp_components.append({
                        "name": exp.get("name", ""),
                        "category": cat,
                        "monthly_amount": round(effective_monthly, 2),
                        "formula": f"${monthly_base:,.0f}/mo{increase_note}",
                    })
            else:
                # Fallback to existing monthly data if no expense schedule configured
                ex = existing_monthly.get((year, month), {})
                office_expense   = float(ex.get("office_expense", 0) or 0)
                legal_accounting = float(ex.get("legal_accounting", 0) or 0)
                software         = float(ex.get("software", 0) or 0)
                direct_expenses  = float(ex.get("direct_expenses", 0) or 0)
                exp_components = [
                    {"name": "Office Expense", "category": "office_expense", "monthly_amount": round(office_expense, 2), "formula": "manual"},
                    {"name": "Legal & Accounting", "category": "legal_accounting", "monthly_amount": round(legal_accounting, 2), "formula": "manual"},
                    {"name": "Software", "category": "software", "monthly_amount": round(software, 2), "formula": "manual"},
                ]

            total_opex = tech_comp + exec_comp + ga_comp + sales_comp + office_expense + legal_accounting + software
            gross_margin = total_revenue - direct_expenses
            ebitda = gross_margin - total_opex

            monthly_data.append({
                "year": year,
                "month": month,
                "month_name": MONTH_NAMES[month - 1],
                "rd_contract_revenue": round(rd_rev, 2),
                "portfolio_revenue": round(portfolio_rev, 2),
                "maintenance_revenue": round(maintenance_rev, 2),
                "grant_revenue": round(grant_rev, 2),
                "consumable_revenue": round(consumable_rev, 2),
                "licensing_revenue": round(licensing_rev, 2),
                "total_revenue": round(total_revenue, 2),
                "tech_comp": round(tech_comp, 2),
                "exec_comp": round(exec_comp, 2),
                "ga_comp": round(ga_comp, 2),
                "sales_comp": round(sales_comp, 2),
                "office_expense": round(office_expense, 2),
                "legal_accounting": round(legal_accounting, 2),
                "software": round(software, 2),
                "direct_expenses": round(direct_expenses, 2),
                "gross_margin": round(gross_margin, 2),
                "total_opex": round(total_opex, 2),
                "ebitda": round(ebitda, 2),
            })

            audit_data.append({
                "year": year,
                "month": month,
                "revenue_components": rev_components,
                "fund_components": fund_components,
                "headcount_components": hc_components,
                "expense_components": exp_components,
            })

    # Roll up annual
    annual_map: dict[int, dict] = {}
    fields_to_sum = [
        "rd_contract_revenue","portfolio_revenue","maintenance_revenue","grant_revenue",
        "consumable_revenue","licensing_revenue",
        "total_revenue","tech_comp","exec_comp","ga_comp","sales_comp",
        "office_expense","legal_accounting","software","direct_expenses",
        "gross_margin","total_opex","ebitda",
    ]
    for m in monthly_data:
        yr = m["year"]
        if yr not in annual_map:
            annual_map[yr] = {"year": yr, "revenue": 0.0, **{f: 0.0 for f in fields_to_sum}}
        a = annual_map[yr]
        for f in fields_to_sum:
            a[f] = round(a[f] + m.get(f, 0), 2)
        a["revenue"] = a["total_revenue"]

    annual_data = sorted(annual_map.values(), key=lambda x: x["year"])
    return {"monthly_data": monthly_data, "annual_data": annual_data, "audit_data": audit_data}


def _snapshot_model(model: dict) -> dict:
    """Serialize a model row to a JSON-safe dict."""
    import json as _json
    snap = dict(model)
    for k, v in snap.items():
        if hasattr(v, "isoformat"):
            snap[k] = v.isoformat()
        elif hasattr(v, "__str__") and not isinstance(v, (str, int, float, bool, list, dict, type(None))):
            snap[k] = str(v)
    return snap


def _next_version_number(cur, bump: str = "patch") -> str:
    cur.execute(
        "SELECT version_number FROM fpa_model_versions WHERE version_number IS NOT NULL ORDER BY created_at DESC LIMIT 1"
    )
    row = cur.fetchone()
    if row and row["version_number"]:
        try:
            parts = row["version_number"].split(".")
            major, minor, patch = int(parts[0]), int(parts[1]), int(parts[2])
        except Exception:
            major, minor, patch = 1, 0, 0
    else:
        major, minor, patch = 1, 0, -1  # first save will become 1.0.0

    if bump == "major":
        major += 1; minor = 0; patch = 0
    elif bump == "minor":
        minor += 1; patch = 0
    else:
        patch += 1

    return f"{major}.{minor}.{patch}"


def _save_version(cur, model: dict, created_by: str | None, change_summary: str | None,
                  scenario_name: str | None, audit_data: list, bump: str = "patch") -> str:
    import json as _json
    snap = _snapshot_model(model)
    version_number = _next_version_number(cur, bump)
    cur.execute(
        """INSERT INTO fpa_model_versions
             (created_by, change_summary, scenario_name, model_snapshot, audit_data, version_number)
           VALUES (%s, %s, %s, %s, %s, %s)
           RETURNING version_id""",
        (created_by, change_summary, scenario_name, _json.dumps(snap), _json.dumps(audit_data), version_number),
    )
    return str(cur.fetchone()["version_id"])


# ── Version history endpoints ─────────────────────────────────────────────────

@router.get("/model/history")
def list_model_history(request: Request, limit: int = 50):
    """Return list of saved model versions (without full snapshot). Admin only."""
    require_admin(request)
    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute(
            """SELECT version_id, created_at, created_by, change_summary, scenario_name, is_default, version_number
               FROM fpa_model_versions ORDER BY created_at DESC LIMIT %s""",
            (limit,),
        )
        rows = cur.fetchall()
        cur.close()
    finally:
        conn.close()
    return [
        {
            **dict(r),
            "version_id": str(r["version_id"]),
            "created_at": r["created_at"].isoformat() if r["created_at"] else None,
        }
        for r in rows
    ]


@router.get("/model/history/{version_id}")
def get_model_version(version_id: str, request: Request):
    """Return full snapshot for a specific version. Admin only."""
    require_admin(request)
    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute("SELECT * FROM fpa_model_versions WHERE version_id = %s", (version_id,))
        row = cur.fetchone()
        cur.close()
    finally:
        conn.close()
    if not row:
        raise HTTPException(status_code=404, detail="Version not found")
    d = dict(row)
    d["version_id"] = str(d["version_id"])
    d["created_at"] = d["created_at"].isoformat() if d["created_at"] else None
    return d


class RestoreBody(BaseModel):
    change_summary: str | None = None
    scenario_name: str | None = None


@router.post("/model/history/{version_id}/restore")
def restore_model_version(version_id: str, body: RestoreBody, request: Request):
    """Restore a historical version as the active model. Admin only."""
    import json as _json
    user = require_admin(request)

    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute("SELECT model_snapshot FROM fpa_model_versions WHERE version_id = %s", (version_id,))
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Version not found")
        snap = row["model_snapshot"]
        if isinstance(snap, str):
            snap = _json.loads(snap)

        # Snapshot current state as a version before overwriting
        current = _fetch_active_model(cur)
        if current:
            _save_version(cur, current, user["email"],
                         f"Auto-saved before restore of {version_id[:8]}…", None, [], bump="patch")

        # Overwrite active model with snapshot data
        fields_to_restore = [
            "exit_multiple","exit_year","start_year","working_capital","starting_valuation",
            "original_equity","non_cash_equity","financing_charges_pct","sba_rate",
            "bank_debt_rate","mezz_rate","hy_rate","rd_contract_avg_value",
            "rd_contract_duration_months","portfolio_contract_avg_value",
            "portfolio_contract_duration_months","annual_data","monthly_data",
            "equity_rounds","notes","contract_pipeline","headcount_schedule","expense_schedule","funding_schedule",
        ]
        set_parts = []
        vals = []
        for f in fields_to_restore:
            if f in snap and snap[f] is not None:
                set_parts.append(f"{f} = %s")
                v = snap[f]
                if isinstance(v, (list, dict)):
                    v = _json.dumps(v)
                vals.append(v)

        if set_parts:
            vals.append(body.change_summary or f"Restored from version {version_id[:8]}…")
            cur.execute(
                f"UPDATE fpa_model SET {', '.join(set_parts)}, change_summary = %s WHERE is_active = true",
                vals,
            )

        conn.commit()
        model = _fetch_active_model(cur)
        cur.close()
    finally:
        conn.close()
    return dict(model)


@router.post("/model/history/{version_id}/set-default")
def set_default_version(version_id: str, request: Request):
    """Mark a version as default and restore it as the active model. Admin only."""
    import json as _json
    user = require_admin(request)

    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute("SELECT model_snapshot FROM fpa_model_versions WHERE version_id = %s", (version_id,))
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Version not found")
        snap = row["model_snapshot"]
        if isinstance(snap, str):
            snap = _json.loads(snap)

        # Snapshot current state before overwriting
        current = _fetch_active_model(cur)
        if current:
            _save_version(cur, current, user["email"], "Auto-saved before set-default", None, [], bump="patch")

        # Restore snapshot to active model
        fields_to_restore = [
            "exit_multiple","exit_year","start_year","working_capital","starting_valuation",
            "original_equity","non_cash_equity","financing_charges_pct","sba_rate",
            "bank_debt_rate","mezz_rate","hy_rate","rd_contract_avg_value",
            "rd_contract_duration_months","portfolio_contract_avg_value",
            "portfolio_contract_duration_months","annual_data","monthly_data",
            "equity_rounds","notes","contract_pipeline","headcount_schedule","expense_schedule","funding_schedule",
        ]
        set_parts, vals = [], []
        for f in fields_to_restore:
            if f in snap and snap[f] is not None:
                set_parts.append(f"{f} = %s")
                v = snap[f]
                if isinstance(v, (list, dict)):
                    v = _json.dumps(v)
                vals.append(v)
        if set_parts:
            vals.append(f"Set as default: version {version_id[:8]}…")
            cur.execute(
                f"UPDATE fpa_model SET {', '.join(set_parts)}, change_summary = %s WHERE is_active = true",
                vals,
            )

        # Mark this version as default, clear others
        cur.execute("UPDATE fpa_model_versions SET is_default = false")
        cur.execute("UPDATE fpa_model_versions SET is_default = true WHERE version_id = %s", (version_id,))

        conn.commit()
        model = _fetch_active_model(cur)
        cur.close()
    finally:
        conn.close()
    return dict(model)


@router.post("/model/save-snapshot")
def save_snapshot(request: Request, body: RestoreBody):
    """Save the current active model state as a named version. Admin only."""
    user = require_admin(request)
    conn = get_conn()
    try:
        cur = conn.cursor()
        model = _fetch_active_model(cur)
        if not model:
            raise HTTPException(status_code=404, detail="No active model found")
        version_id = _save_version(
            cur, dict(model),
            user["email"],
            body.change_summary or body.scenario_name or "Manual snapshot",
            body.scenario_name,
            [],
            bump="minor" if body.scenario_name else "patch",
        )
        conn.commit()
        cur.close()
    finally:
        conn.close()
    return {"version_id": version_id, "ok": True}


# ── Calculation engine endpoint ───────────────────────────────────────────────

@router.post("/model/recalculate")
def recalculate_model(request: Request):
    """Run the driver-based engine on the active model, persist results, and return. Admin only."""
    import json as _json
    require_admin(request)
    conn = get_conn()
    try:
        cur = conn.cursor()
        model = _fetch_active_model(cur)
        if not model:
            raise HTTPException(status_code=404, detail="No active model found")
        result = _run_engine(dict(model))
        cur.execute(
            "UPDATE fpa_model SET annual_data = %s, monthly_data = %s WHERE is_active = true",
            (_json.dumps(result["annual_data"]), _json.dumps(result["monthly_data"])),
        )
        conn.commit()
        cur.close()
    finally:
        conn.close()
    return result


class CommitDriversBody(BaseModel):
    contract_pipeline: list[dict[str, Any]] | None = None
    headcount_schedule: list[dict[str, Any]] | None = None
    expense_schedule: list[dict[str, Any]] | None = None
    funding_schedule: list[dict[str, Any]] | None = None
    change_summary: str | None = None
    scenario_name: str | None = None


@router.post("/model/commit-drivers")
def commit_drivers(body: CommitDriversBody, request: Request):
    """Save driver sheets, run engine, update active model, and save a version entry. Admin only."""
    import json as _json
    user = require_admin(request)

    conn = get_conn()
    try:
        cur = conn.cursor()
        # Update driver columns
        update_parts = []
        update_vals = []
        if body.contract_pipeline is not None:
            update_parts.append("contract_pipeline = %s")
            update_vals.append(_json.dumps(body.contract_pipeline))
        if body.headcount_schedule is not None:
            update_parts.append("headcount_schedule = %s")
            update_vals.append(_json.dumps(body.headcount_schedule))
        if body.expense_schedule is not None:
            update_parts.append("expense_schedule = %s")
            update_vals.append(_json.dumps(body.expense_schedule))
        if body.funding_schedule is not None:
            update_parts.append("funding_schedule = %s")
            update_vals.append(_json.dumps(body.funding_schedule))
        if body.change_summary is not None:
            update_parts.append("change_summary = %s")
            update_vals.append(body.change_summary)
        if update_parts:
            cur.execute(
                f"UPDATE fpa_model SET {', '.join(update_parts)} WHERE is_active = true",
                update_vals,
            )
            conn.commit()

        model = _fetch_active_model(cur)
        if not model:
            raise HTTPException(status_code=404, detail="No active model found")

        # Run engine
        engine_result = _run_engine(dict(model))

        # Persist computed P&L back to model
        cur.execute(
            "UPDATE fpa_model SET annual_data = %s, monthly_data = %s WHERE is_active = true",
            (_json.dumps(engine_result["annual_data"]), _json.dumps(engine_result["monthly_data"])),
        )

        # Save version entry
        model_after = _fetch_active_model(cur)
        version_id = _save_version(
            cur, dict(model_after),
            user["email"],
            body.change_summary or "Driver update",
            body.scenario_name,
            engine_result["audit_data"],
            bump="minor",
        )
        conn.commit()
        cur.close()
    finally:
        conn.close()

    return {
        "version_id": version_id,
        "annual_data": engine_result["annual_data"],
        "monthly_data": engine_result["monthly_data"],
        "audit_data": engine_result["audit_data"],
    }


@router.get("/model/audit")
def get_audit(request: Request, version_id: str | None = None):
    """Return audit data for a version or the most recent version. Admin only."""
    require_admin(request)
    conn = get_conn()
    try:
        cur = conn.cursor()
        if version_id:
            cur.execute(
                "SELECT audit_data, created_at, change_summary FROM fpa_model_versions WHERE version_id = %s",
                (version_id,),
            )
        else:
            cur.execute(
                "SELECT audit_data, created_at, change_summary FROM fpa_model_versions ORDER BY created_at DESC LIMIT 1"
            )
        row = cur.fetchone()
        cur.close()
    finally:
        conn.close()
    if not row:
        return {"audit_data": [], "created_at": None, "change_summary": None}
    return {
        "audit_data": row["audit_data"],
        "created_at": row["created_at"].isoformat() if row["created_at"] else None,
        "change_summary": row["change_summary"],
    }


# ── Scenario load endpoint ────────────────────────────────────────────────────

@router.post("/scenarios/{scenario_id}/load")
def load_scenario(scenario_id: str, request: Request):
    """Load a saved scenario as the active model state. Admin only."""
    import json as _json
    user = require_admin(request)

    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute("SELECT * FROM fpa_scenarios WHERE scenario_id = %s", (scenario_id,))
        scen = cur.fetchone()
        if not scen:
            raise HTTPException(status_code=404, detail="Scenario not found")

        snap = scen["model_snapshot"]
        if isinstance(snap, str):
            snap = _json.loads(snap)

        # Save current state before overwriting
        current = _fetch_active_model(cur)
        if current:
            _save_version(cur, current, user["email"],
                         f"Auto-saved before loading scenario '{scen['name']}'", None, [], bump="patch")

        fields_to_restore = [
            "exit_multiple","exit_year","start_year","working_capital","starting_valuation",
            "original_equity","non_cash_equity","financing_charges_pct","sba_rate",
            "bank_debt_rate","mezz_rate","hy_rate","rd_contract_avg_value",
            "rd_contract_duration_months","portfolio_contract_avg_value",
            "portfolio_contract_duration_months","annual_data","monthly_data",
            "equity_rounds","notes","contract_pipeline","headcount_schedule","expense_schedule","funding_schedule",
        ]
        set_parts = []
        vals = []
        for f in fields_to_restore:
            if f in snap and snap[f] is not None:
                set_parts.append(f"{f} = %s")
                v = snap[f]
                if isinstance(v, (list, dict)):
                    v = _json.dumps(v)
                vals.append(v)
        if set_parts:
            vals.append(f"Loaded scenario: {scen['name']}")
            cur.execute(
                f"UPDATE fpa_model SET {', '.join(set_parts)}, change_summary = %s WHERE is_active = true",
                vals,
            )
        conn.commit()
        model = _fetch_active_model(cur)
        cur.close()
    finally:
        conn.close()
    return dict(model)


# ── Clean Excel export (v2) ───────────────────────────────────────────────────

@router.get("/model/export-v2")
def export_model_v2(request: Request):
    """Generate a clean Excel workbook from the active model. Admin only."""
    require_admin(request)
    conn = get_conn()
    try:
        cur = conn.cursor()
        model = _fetch_active_model(cur)
        # Get most recent version for audit data
        cur.execute(
            "SELECT audit_data, created_at, change_summary FROM fpa_model_versions ORDER BY created_at DESC LIMIT 1"
        )
        ver_row = cur.fetchone()
        cur.close()
    finally:
        conn.close()

    if not model:
        raise HTTPException(status_code=404, detail="No active model found")

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    m = dict(model)
    annual_data = m.get("annual_data") or []
    monthly_data = m.get("monthly_data") or []
    contracts = m.get("contract_pipeline") or []
    headcount = m.get("headcount_schedule") or []
    expense_schedule = m.get("expense_schedule") or []
    funding_schedule = m.get("funding_schedule") or []
    audit_data = (ver_row["audit_data"] if ver_row else []) or []

    wb = openpyxl.Workbook()

    # Palette
    NAVY    = "1F3864"
    BLUE    = "2563EB"
    GREEN   = "16A34A"
    RED     = "DC2626"
    AMBER   = "D97706"
    WHITE   = "FFFFFF"
    LGRAY   = "F3F4F6"
    MGRAY   = "E5E7EB"
    BLACK   = "111827"
    TEAL    = "0891B2"

    thin = Side(border_style="thin", color="D1D5DB")
    med  = Side(border_style="medium", color="9CA3AF")

    def _hdr(cell, text, bg=NAVY, fg=WHITE, bold=True, size=11, align="center"):
        cell.value = text
        cell.font = Font(bold=bold, color=fg, size=size)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(border_style="medium", color=bg))

    def _val(cell, v, fmt=None, color=BLACK, bold=False):
        cell.value = v
        cell.font = Font(color=color, bold=bold, size=10)
        cell.alignment = Alignment(horizontal="right" if isinstance(v, (int, float)) else "left", vertical="center")
        if fmt:
            cell.number_format = fmt
        cell.border = Border(bottom=thin)

    def _lbl(cell, text, bold=False, color=BLACK):
        cell.value = text
        cell.font = Font(bold=bold, color=color, size=10)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = Border(bottom=thin)

    MONEY  = '$#,##0_);[Red]($#,##0)'
    MONEYK = '$#,##0,"K"_);[Red]($#,##0,"K")'
    PCT    = '0.0%'
    DATE_FMT = 'YYYY-MM-DD'

    years = [d["year"] for d in annual_data if isinstance(d, dict)]

    # ──────────────────────────────────────────────────────────────────────────
    # SHEET 1: Cover
    # ──────────────────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Cover"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value = "Collective ERP — Financial Model"
    c.font = Font(bold=True, size=18, color=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center")

    info_rows = [
        ("Generated", __import__("datetime").datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Change Summary", m.get("change_summary") or "—"),
        ("Start Year", m.get("start_year", 2025)),
        ("Exit Year", m.get("exit_year", 10)),
        ("Exit Multiple", f"{m.get('exit_multiple', 4.0)}×"),
        ("Starting Valuation", m.get("starting_valuation", 0)),
        ("Contracts in Pipeline", len(contracts)),
        ("Headcount (current)", len(headcount)),
    ]
    for i, (lbl, val) in enumerate(info_rows, start=3):
        ws.cell(i, 1).value = lbl
        ws.cell(i, 1).font = Font(bold=True, color="6B7280", size=10)
        ws.cell(i, 2).value = val
        ws.cell(i, 2).font = Font(color=BLACK, size=10)
        if isinstance(val, (int, float)) and i in (8,):
            ws.cell(i, 2).number_format = MONEY

    ws.cell(12, 1).value = "Sheets"
    ws.cell(12, 1).font = Font(bold=True, color=NAVY, size=11)
    sheets_list = [
        ("P&L Annual", "Annual P&L by year, all revenue streams and expenses"),
        ("P&L Monthly", "Month-by-month P&L detail"),
        ("Contract Pipeline", "All contracts with probability-weighted revenue"),
        ("Headcount", "Roster with fully-loaded monthly costs"),
        ("Assumptions", "Scalar model assumptions and financing rates"),
        ("Equity Rounds", "Cap table and equity financing rounds"),
        ("Audit Log", "Formula trace for each monthly revenue and cost cell"),
    ]
    for i, (sh, desc) in enumerate(sheets_list, start=13):
        ws.cell(i, 1).value = sh
        ws.cell(i, 1).font = Font(bold=True, color=BLUE, size=10)
        ws.cell(i, 2).value = desc
        ws.cell(i, 2).font = Font(color="4B5563", size=10)

    # ──────────────────────────────────────────────────────────────────────────
    # SHEET 2: P&L Annual
    # ──────────────────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("P&L Annual")
    ws2.column_dimensions["A"].width = 32
    for j, yr in enumerate(years, 2):
        ws2.column_dimensions[get_column_letter(j)].width = 14

    ws2.merge_cells(f"A1:{get_column_letter(max(2, len(years)+1))}1")
    c = ws2["A1"]
    c.value = "Annual P&L Summary ($)"
    c.font = Font(bold=True, size=13, color=NAVY)
    c.fill = PatternFill("solid", fgColor=LGRAY)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[1].height = 22

    _hdr(ws2["A2"], "Line Item", align="left")
    for j, yr in enumerate(years, 2):
        _hdr(ws2.cell(2, j), yr)

    rows_def = [
        ("Revenue", "total_revenue", MONEY, NAVY, True),
        ("  R&D Contracts", "rd_contract_revenue", MONEY, "374151", False),
        ("  Portfolio", "portfolio_revenue", MONEY, "374151", False),
        ("  Maintenance", "maintenance_revenue", MONEY, "374151", False),
        ("  Grants", "grant_revenue", MONEY, "374151", False),
        ("Direct Expenses", "direct_expenses", MONEY, RED, False),
        ("Gross Margin", "gross_margin", MONEY, GREEN, True),
        (None, None, None, None, None),
        ("  Technical Comp", "tech_comp", MONEY, "374151", False),
        ("  Executive Comp", "exec_comp", MONEY, "374151", False),
        ("  G&A Comp", "ga_comp", MONEY, "374151", False),
        ("  Sales Comp", "sales_comp", MONEY, "374151", False),
        ("  Office Expense", "office_expense", MONEY, "374151", False),
        ("  Legal & Accounting", "legal_accounting", MONEY, "374151", False),
        ("  Software", "software", MONEY, "374151", False),
        ("Total OpEx", "total_opex", MONEY, RED, True),
        (None, None, None, None, None),
        ("EBITDA", "ebitda", MONEY, NAVY, True),
        ("EBITDA Margin %", "_ebitda_pct", PCT, TEAL, False),
    ]

    annual_by_year = {d["year"]: d for d in annual_data if isinstance(d, dict)}
    row_i = 3
    for label, field, fmt, color, bold in rows_def:
        if label is None:
            row_i += 1
            continue
        _lbl(ws2.cell(row_i, 1), label, bold=bold, color=color)
        if bold:
            ws2.cell(row_i, 1).fill = PatternFill("solid", fgColor=LGRAY)
        for j, yr in enumerate(years, 2):
            d = annual_by_year.get(yr, {})
            if field == "_ebitda_pct":
                rev = d.get("total_revenue", 1) or 1
                v = d.get("ebitda", 0) / rev
            else:
                v = d.get(field, 0)
            c2 = ws2.cell(row_i, j)
            _val(c2, v, fmt, color, bold)
            if bold:
                c2.fill = PatternFill("solid", fgColor=LGRAY)
        row_i += 1

    # ──────────────────────────────────────────────────────────────────────────
    # SHEET 3: P&L Monthly  (rows = line items, columns = months)
    # ──────────────────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("P&L Monthly")

    # Build ordered month list from monthly_data
    month_order = sorted(
        {(m["year"], m["month"]) for m in monthly_data},
        key=lambda x: (x[0], x[1]),
    )
    # Lookup: (year, month) → row dict
    mo_lookup: dict[tuple, dict] = {(m["year"], m["month"]): m for m in monthly_data}

    # Row definitions: (display label, field, is_bold, is_subtotal, indent)
    monthly_rows_def = [
        ("Total Revenue",          "total_revenue",         True,  True,  False),
        ("  R&D Contracts",        "rd_contract_revenue",   False, False, True),
        ("  Portfolio",            "portfolio_revenue",     False, False, True),
        ("  Maintenance",          "maintenance_revenue",   False, False, True),
        ("  Grants",               "grant_revenue",         False, False, True),
        ("Direct Expenses",        "direct_expenses",       False, False, False),
        ("Gross Margin",           "gross_margin",          True,  True,  False),
        (None,                     None,                    None,  None,  None),
        ("  Technical Comp",       "tech_comp",             False, False, True),
        ("  Executive Comp",       "exec_comp",             False, False, True),
        ("  G&A Comp",             "ga_comp",               False, False, True),
        ("  Sales Comp",           "sales_comp",            False, False, True),
        ("  Office Expense",       "office_expense",        False, False, True),
        ("  Legal & Accounting",   "legal_accounting",      False, False, True),
        ("  Software",             "software",              False, False, True),
        ("Total OpEx",             "total_opex",            True,  True,  False),
        (None,                     None,                    None,  None,  None),
        ("EBITDA",                 "ebitda",                True,  True,  False),
    ]

    # Header row 1: year labels spanning 12 months each
    ws3.column_dimensions["A"].width = 26
    ws3.cell(1, 1).value = "P&L Monthly ($)"
    ws3.cell(1, 1).font = Font(bold=True, size=12, color=NAVY)
    ws3.cell(2, 1).value = "Line Item"
    ws3.cell(2, 1).font = Font(bold=True, color=WHITE, size=10)
    ws3.cell(2, 1).fill = PatternFill("solid", fgColor=NAVY)
    ws3.cell(2, 1).alignment = Alignment(horizontal="left", vertical="center")

    col = 2
    for yr, mo in month_order:
        # Year header (row 1): spans Jan of year to Dec
        if mo == 1:
            yr_col_start = col
            yr_col_end = col + 11
            ws3.merge_cells(start_row=1, start_column=yr_col_start, end_row=1, end_column=yr_col_end)
            yc = ws3.cell(1, yr_col_start)
            yc.value = yr
            yc.font = Font(bold=True, color=WHITE, size=11)
            yc.fill = PatternFill("solid", fgColor=NAVY)
            yc.alignment = Alignment(horizontal="center", vertical="center")
        # Month header (row 2)
        mc = ws3.cell(2, col)
        mo_abbr = MONTH_NAMES[mo - 1][:3]
        mc.value = mo_abbr
        mc.font = Font(bold=True, color=WHITE, size=9)
        mc.fill = PatternFill("solid", fgColor="374151")
        mc.alignment = Alignment(horizontal="center", vertical="center")
        ws3.column_dimensions[get_column_letter(col)].width = 10
        col += 1

    ws3.row_dimensions[1].height = 18
    ws3.row_dimensions[2].height = 16
    ws3.freeze_panes = "B3"

    # Data rows
    data_row = 3
    for label, field, bold, is_subtotal, indent in monthly_rows_def:
        if label is None:
            data_row += 1
            continue
        # Row label
        lc = ws3.cell(data_row, 1)
        lc.value = label
        lc.font = Font(bold=bold, color=NAVY if is_subtotal else ("374151" if not indent else "6B7280"), size=10)
        lc.alignment = Alignment(horizontal="left", vertical="center")
        if is_subtotal:
            lc.fill = PatternFill("solid", fgColor=LGRAY)
        lc.border = Border(bottom=thin)

        col = 2
        for yr, mo in month_order:
            mo_dict = mo_lookup.get((yr, mo), {})
            val = mo_dict.get(field, 0) or 0
            vc = ws3.cell(data_row, col)
            vc.value = val
            vc.number_format = MONEY
            vc.alignment = Alignment(horizontal="right", vertical="center")
            vc.border = Border(bottom=thin)
            if is_subtotal:
                vc.fill = PatternFill("solid", fgColor=LGRAY)
                vc.font = Font(bold=True, size=9,
                    color=(GREEN if field == "ebitda" and val >= 0
                           else RED if field in ("total_opex","direct_expenses") or (field == "ebitda" and val < 0)
                           else NAVY))
            else:
                vc.font = Font(size=9, color="374151")
            col += 1
        data_row += 1

    # ──────────────────────────────────────────────────────────────────────────
    # SHEET 4: Contract Pipeline
    # ──────────────────────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Contract Pipeline")
    cp_headers = ["Name","Type","Client","Start Date","Duration (mo)","Total Value","Probability","Wtd Value","Status","Notes"]
    col_widths4 = [24, 14, 20, 14, 14, 16, 12, 16, 12, 28]
    for j, (h, w) in enumerate(zip(cp_headers, col_widths4), 1):
        _hdr(ws4.cell(1, j), h)
        ws4.column_dimensions[get_column_letter(j)].width = w

    for i, c in enumerate(contracts, 2):
        tv = float(c.get("total_value") or 0)
        prob = float(c.get("probability") or 1.0)
        ws4.cell(i, 1).value = c.get("name", "")
        ws4.cell(i, 2).value = c.get("type", "")
        ws4.cell(i, 3).value = c.get("client", "")
        ws4.cell(i, 4).value = (c.get("start_date") or "")[:10]
        ws4.cell(i, 5).value = c.get("duration_months")
        ws4.cell(i, 5).number_format = "0"
        ws4.cell(i, 6).value = tv
        ws4.cell(i, 6).number_format = MONEY
        ws4.cell(i, 7).value = prob
        ws4.cell(i, 7).number_format = PCT
        ws4.cell(i, 8).value = round(tv * prob, 2)
        ws4.cell(i, 8).number_format = MONEY
        ws4.cell(i, 9).value = c.get("status", "")
        ws4.cell(i, 10).value = c.get("notes", "")
        for j in range(1, 11):
            ws4.cell(i, j).border = Border(bottom=thin)
            ws4.cell(i, j).alignment = Alignment(horizontal="right" if j in (5,6,7,8) else "left", vertical="center")
            ws4.cell(i, j).font = Font(size=10)
        ws4.cell(i, 8).font = Font(bold=True, color=GREEN, size=10)

    # Totals row
    tr = len(contracts) + 2
    ws4.cell(tr, 1).value = "TOTAL"
    ws4.cell(tr, 1).font = Font(bold=True, size=10, color=NAVY)
    ws4.cell(tr, 6).value = sum(float(c.get("total_value") or 0) for c in contracts)
    ws4.cell(tr, 6).number_format = MONEY
    ws4.cell(tr, 6).font = Font(bold=True, size=10, color=NAVY)
    ws4.cell(tr, 8).value = sum(float(c.get("total_value") or 0) * float(c.get("probability") or 1) for c in contracts)
    ws4.cell(tr, 8).number_format = MONEY
    ws4.cell(tr, 8).font = Font(bold=True, size=10, color=GREEN)

    # ──────────────────────────────────────────────────────────────────────────
    # SHEET 5: Headcount
    # ──────────────────────────────────────────────────────────────────────────
    ws5 = wb.create_sheet("Headcount")
    hc_headers = ["Role","Department","Start Date","End Date","Annual Salary","Benefits %","Raise %/yr","Monthly Yr1 (Loaded)","Notes"]
    col_widths5 = [28, 14, 14, 14, 18, 12, 12, 20, 28]
    for j, (h, w) in enumerate(zip(hc_headers, col_widths5), 1):
        _hdr(ws5.cell(1, j), h)
        ws5.column_dimensions[get_column_letter(j)].width = w

    for i, e in enumerate(headcount, 2):
        sal = float(e.get("annual_salary") or 0)
        bp  = float(e.get("benefits_pct") or 0.25)
        rp  = float(e.get("annual_raise_pct") or 0)
        monthly_loaded = round(sal / 12 * (1 + bp), 2)
        ws5.cell(i, 1).value = e.get("role", "")
        ws5.cell(i, 2).value = e.get("department", "")
        ws5.cell(i, 3).value = (e.get("start_date") or "")[:10]
        ws5.cell(i, 4).value = (e.get("end_date") or "")[:10]
        ws5.cell(i, 5).value = sal
        ws5.cell(i, 5).number_format = MONEY
        ws5.cell(i, 6).value = bp
        ws5.cell(i, 6).number_format = PCT
        ws5.cell(i, 7).value = rp
        ws5.cell(i, 7).number_format = PCT
        ws5.cell(i, 8).value = monthly_loaded
        ws5.cell(i, 8).number_format = MONEY
        ws5.cell(i, 8).font = Font(bold=True, color=TEAL, size=10)
        ws5.cell(i, 9).value = e.get("notes", "")
        for j in range(1, 10):
            ws5.cell(i, j).border = Border(bottom=thin)
            ws5.cell(i, j).alignment = Alignment(horizontal="right" if j in (5,6,7,8) else "left", vertical="center")

    # Totals
    htr = len(headcount) + 2
    ws5.cell(htr, 1).value = f"TOTAL ({len(headcount)} employees)"
    ws5.cell(htr, 1).font = Font(bold=True, size=10, color=NAVY)
    ws5.cell(htr, 5).value = sum(float(e.get("annual_salary") or 0) for e in headcount)
    ws5.cell(htr, 5).number_format = MONEY
    ws5.cell(htr, 5).font = Font(bold=True, color=NAVY, size=10)
    ws5.cell(htr, 8).value = sum(
        float(e.get("annual_salary") or 0) / 12 * (1 + float(e.get("benefits_pct") or 0.25))
        for e in headcount
    )
    ws5.cell(htr, 8).number_format = MONEY
    ws5.cell(htr, 8).font = Font(bold=True, color=GREEN, size=10)

    # ──────────────────────────────────────────────────────────────────────────
    # SHEET 6: Expense Schedule
    # ──────────────────────────────────────────────────────────────────────────
    CATEGORY_LABELS = {
        "office_expense": "Office / Rent & Utilities",
        "legal_accounting": "Legal & Accounting",
        "software": "Software & Apps",
        "direct_expenses": "Lab Consumables / Direct",
    }
    ws6 = wb.create_sheet("Expense Schedule")
    exp_headers = ["Name", "Category", "Start Date", "End Date", "Monthly Amount", "Annual Increase %", "Notes"]
    col_widths6 = [32, 26, 14, 14, 18, 18, 30]
    for j, (h, w) in enumerate(zip(exp_headers, col_widths6), 1):
        _hdr(ws6.cell(1, j), h)
        ws6.column_dimensions[get_column_letter(j)].width = w

    for i, e in enumerate(expense_schedule, 2):
        cat = e.get("category", "office_expense")
        amt = float(e.get("monthly_amount") or 0)
        inc = float(e.get("annual_increase_pct") or 0)
        ws6.cell(i, 1).value = e.get("name", "")
        ws6.cell(i, 2).value = CATEGORY_LABELS.get(cat, cat)
        ws6.cell(i, 3).value = (e.get("start_date") or "")[:10]
        ws6.cell(i, 4).value = (e.get("end_date") or "")[:10]
        ws6.cell(i, 5).value = amt
        ws6.cell(i, 5).number_format = MONEY
        ws6.cell(i, 6).value = inc
        ws6.cell(i, 6).number_format = PCT
        ws6.cell(i, 7).value = e.get("notes", "")
        for j in range(1, 8):
            ws6.cell(i, j).border = Border(bottom=thin)
            ws6.cell(i, j).alignment = Alignment(horizontal="right" if j in (5, 6) else "left", vertical="center")

    if expense_schedule:
        etr = len(expense_schedule) + 2
        ws6.cell(etr, 1).value = f"TOTAL ({len(expense_schedule)} items)"
        ws6.cell(etr, 1).font = Font(bold=True, size=10, color=NAVY)
        ws6.cell(etr, 5).value = sum(float(e.get("monthly_amount") or 0) for e in expense_schedule)
        ws6.cell(etr, 5).number_format = MONEY
        ws6.cell(etr, 5).font = Font(bold=True, color=GREEN, size=10)

    # ──────────────────────────────────────────────────────────────────────────
    # SHEET 7: Assumptions (renumbered)
    # ──────────────────────────────────────────────────────────────────────────
    ws6 = wb.create_sheet("Assumptions")
    ws6.column_dimensions["A"].width = 36
    ws6.column_dimensions["B"].width = 20
    ws6.merge_cells("A1:B1")
    ws6["A1"].value = "Model Assumptions"
    ws6["A1"].font = Font(bold=True, size=13, color=NAVY)

    assumption_rows = [
        ("EXIT ASSUMPTIONS", None, None),
        ("Exit Multiple of EBITDA", m.get("exit_multiple"), '0.0"x"'),
        ("Exit Year (from start)", m.get("exit_year"), "0"),
        ("Starting Valuation", m.get("starting_valuation"), MONEY),
        ("Original Equity Investment", m.get("original_equity"), MONEY),
        ("Non-Cash Equity Value", m.get("non_cash_equity"), MONEY),
        ("Working Capital", m.get("working_capital"), MONEY),
        (None, None, None),
        ("DEBT FINANCING RATES", None, None),
        ("SBA Loan Rate", m.get("sba_rate"), PCT),
        ("Bank Debt Rate", m.get("bank_debt_rate"), PCT),
        ("Mezzanine Rate", m.get("mezz_rate"), PCT),
        ("High Yield Rate", m.get("hy_rate"), PCT),
        (None, None, None),
        ("CONTRACT PARAMETERS", None, None),
        ("R&D Contract Avg Value", m.get("rd_contract_avg_value"), MONEY),
        ("R&D Contract Duration (mo)", m.get("rd_contract_duration_months"), "0"),
        ("Portfolio Contract Avg Value", m.get("portfolio_contract_avg_value"), MONEY),
        ("Portfolio Contract Duration (mo)", m.get("portfolio_contract_duration_months"), "0"),
    ]
    row_i = 3
    for label, val, fmt in assumption_rows:
        if label is None:
            row_i += 1
            continue
        if val is None:
            ws6.cell(row_i, 1).value = label
            ws6.cell(row_i, 1).font = Font(bold=True, color=NAVY, size=11)
            ws6.cell(row_i, 1).fill = PatternFill("solid", fgColor=LGRAY)
            ws6.cell(row_i, 2).fill = PatternFill("solid", fgColor=LGRAY)
        else:
            ws6.cell(row_i, 1).value = label
            ws6.cell(row_i, 1).font = Font(color="374151", size=10)
            ws6.cell(row_i, 2).value = val
            ws6.cell(row_i, 2).font = Font(color=BLUE, size=10)
            ws6.cell(row_i, 2).fill = PatternFill("solid", fgColor="FFFBEB")
            if fmt:
                ws6.cell(row_i, 2).number_format = fmt
            ws6.cell(row_i, 2).alignment = Alignment(horizontal="right")
        row_i += 1

    # ──────────────────────────────────────────────────────────────────────────
    # SHEET 7: Funding Schedule
    # ──────────────────────────────────────────────────────────────────────────
    FUND_TYPE_LABELS = {
        "equity": "Equity Round", "safe": "SAFE", "convertible_note": "Convertible Note",
        "grant": "Grant", "sbir": "SBIR/STTR", "loan": "Loan / Debt",
    }
    ws7 = wb.create_sheet("Funding Schedule")

    # ── Section A: Dilutive ──
    ws7.cell(1, 1).value = "DILUTIVE FUNDING"
    ws7.cell(1, 1).font = Font(bold=True, size=11, color=NAVY)
    dil_headers = ["Name", "Type", "Close Date", "Amount", "Pre-Money Val.", "Dilution %", "Notes"]
    dil_widths   = [28, 18, 14, 18, 20, 12, 30]
    for j, (h, w) in enumerate(zip(dil_headers, dil_widths), 1):
        _hdr(ws7.cell(2, j), h)
        ws7.column_dimensions[get_column_letter(j)].width = w

    _dilutive = {"equity", "safe", "convertible_note"}
    dil_rows = [f for f in funding_schedule if f.get("type", "equity") in _dilutive]
    # Also pull legacy equity_rounds entries
    for rnd in (m.get("equity_rounds") or []):
        dil_rows.append({
            "name": f"Round {rnd.get('round', '')}",
            "type": "equity",
            "date": rnd.get("date", ""),
            "amount": rnd.get("amount", 0),
            "notes": rnd.get("notes", ""),
        })

    dil_cumulative = 0
    for i, f in enumerate(dil_rows, 3):
        amt = float(f.get("amount") or 0)
        dil_cumulative += amt
        pre = float(f.get("pre_money_valuation") or 0)
        dil_pct = float(f.get("dilution_pct") or (amt / (pre + amt) if pre > 0 else 0))
        ws7.cell(i, 1).value = f.get("name", "")
        ws7.cell(i, 2).value = FUND_TYPE_LABELS.get(f.get("type", "equity"), f.get("type", ""))
        ws7.cell(i, 3).value = (f.get("date") or "")[:10]
        ws7.cell(i, 4).value = amt
        ws7.cell(i, 4).number_format = MONEY
        ws7.cell(i, 5).value = pre if pre > 0 else ""
        if pre > 0:
            ws7.cell(i, 5).number_format = MONEY
        ws7.cell(i, 6).value = dil_pct if dil_pct > 0 else ""
        if dil_pct > 0:
            ws7.cell(i, 6).number_format = PCT
        ws7.cell(i, 7).value = f.get("notes", "")
        for j in range(1, 8):
            ws7.cell(i, j).border = Border(bottom=thin)
            ws7.cell(i, j).alignment = Alignment(horizontal="right" if j in (4,5,6) else "left", vertical="center")
    dil_total_row = len(dil_rows) + 3
    ws7.cell(dil_total_row, 1).value = f"Total Dilutive ({len(dil_rows)} rounds)"
    ws7.cell(dil_total_row, 1).font = Font(bold=True, size=10, color=NAVY)
    ws7.cell(dil_total_row, 4).value = dil_cumulative
    ws7.cell(dil_total_row, 4).number_format = MONEY
    ws7.cell(dil_total_row, 4).font = Font(bold=True, color=GREEN, size=10)

    # ── Section B: Non-Dilutive ──
    nd_start = dil_total_row + 2
    ws7.cell(nd_start, 1).value = "NON-DILUTIVE FUNDING"
    ws7.cell(nd_start, 1).font = Font(bold=True, size=11, color=NAVY)
    nd_headers = ["Name", "Type", "Disbursement", "Start / Date", "End Date", "Amount / Monthly", "Annual Increase", "Notes"]
    nd_widths   = [28, 18, 14, 14, 14, 20, 16, 30]
    for j, (h, w) in enumerate(zip(nd_headers, nd_widths), 1):
        _hdr(ws7.cell(nd_start + 1, j), h)
        ws7.column_dimensions[get_column_letter(j)].width = max(ws7.column_dimensions[get_column_letter(j)].width, w)

    _nondil = {"grant", "sbir", "loan"}
    nd_rows = [f for f in funding_schedule if f.get("type", "equity") in _nondil]
    nd_total = 0
    for i, f in enumerate(nd_rows, nd_start + 2):
        disb = f.get("disbursement", "lump_sum")
        amt = float(f.get("amount") or f.get("monthly_amount") or 0)
        nd_total += float(f.get("amount") or 0)
        ws7.cell(i, 1).value = f.get("name", "")
        ws7.cell(i, 2).value = FUND_TYPE_LABELS.get(f.get("type", ""), f.get("type", ""))
        ws7.cell(i, 3).value = "Monthly" if disb == "monthly" else "Lump Sum"
        ws7.cell(i, 4).value = (f.get("start_date") or f.get("date") or "")[:10]
        ws7.cell(i, 5).value = (f.get("end_date") or "")[:10]
        ws7.cell(i, 6).value = amt
        ws7.cell(i, 6).number_format = MONEY
        inc = float(f.get("annual_increase_pct") or 0)
        ws7.cell(i, 7).value = inc if inc > 0 else ""
        if inc > 0:
            ws7.cell(i, 7).number_format = PCT
        ws7.cell(i, 8).value = f.get("notes", "")
        for j in range(1, 9):
            ws7.cell(i, j).border = Border(bottom=thin)
            ws7.cell(i, j).alignment = Alignment(horizontal="right" if j in (6, 7) else "left", vertical="center")
    nd_total_row = nd_start + 2 + len(nd_rows)
    ws7.cell(nd_total_row, 1).value = f"Total Non-Dilutive ({len(nd_rows)} items)"
    ws7.cell(nd_total_row, 1).font = Font(bold=True, size=10, color=NAVY)
    ws7.cell(nd_total_row, 6).value = nd_total
    ws7.cell(nd_total_row, 6).number_format = MONEY
    ws7.cell(nd_total_row, 6).font = Font(bold=True, color=GREEN, size=10)

    # ──────────────────────────────────────────────────────────────────────────
    # SHEET 8: Audit Log
    # ──────────────────────────────────────────────────────────────────────────
    ws8 = wb.create_sheet("Audit Log")
    audit_headers = ["Year","Month","Category","Item Name","Value ($)","Formula","Source"]
    audit_widths  = [8, 12, 18, 32, 16, 44, 14]
    for j, (h, w) in enumerate(zip(audit_headers, audit_widths), 1):
        _hdr(ws8.cell(1, j), h)
        ws8.column_dimensions[get_column_letter(j)].width = w

    row_i = 2
    for entry in (audit_data or []):
        yr = entry.get("year")
        mo_name = MONTH_NAMES[entry.get("month", 1) - 1][:3] if entry.get("month") else ""
        for comp in entry.get("revenue_components", []):
            ws8.cell(row_i, 1).value = yr
            ws8.cell(row_i, 2).value = mo_name
            ws8.cell(row_i, 3).value = "Revenue"
            ws8.cell(row_i, 4).value = comp.get("name", "")
            ws8.cell(row_i, 5).value = comp.get("monthly_value", 0)
            ws8.cell(row_i, 5).number_format = MONEY
            ws8.cell(row_i, 6).value = comp.get("formula", "")
            ws8.cell(row_i, 7).value = "Contract Pipeline"
            for j in range(1, 8):
                ws8.cell(row_i, j).font = Font(size=9)
                ws8.cell(row_i, j).border = Border(bottom=thin)
            row_i += 1
        for comp in entry.get("fund_components", []):
            ws8.cell(row_i, 1).value = yr
            ws8.cell(row_i, 2).value = mo_name
            ws8.cell(row_i, 3).value = FUND_TYPE_LABELS.get(comp.get("type", ""), comp.get("type", ""))
            ws8.cell(row_i, 4).value = comp.get("name", "")
            ws8.cell(row_i, 5).value = comp.get("monthly_value", 0)
            ws8.cell(row_i, 5).number_format = MONEY
            ws8.cell(row_i, 6).value = comp.get("formula", "")
            ws8.cell(row_i, 7).value = "Funding Schedule"
            for j in range(1, 8):
                ws8.cell(row_i, j).font = Font(size=9)
                ws8.cell(row_i, j).border = Border(bottom=thin)
            row_i += 1
        for comp in entry.get("headcount_components", []):
            ws8.cell(row_i, 1).value = yr
            ws8.cell(row_i, 2).value = mo_name
            ws8.cell(row_i, 3).value = f"Comp ({comp.get('department', '')})"
            ws8.cell(row_i, 4).value = comp.get("role", "")
            ws8.cell(row_i, 5).value = comp.get("monthly_cost", 0)
            ws8.cell(row_i, 5).number_format = MONEY
            ws8.cell(row_i, 6).value = comp.get("formula", "")
            ws8.cell(row_i, 7).value = "Headcount"
            for j in range(1, 8):
                ws8.cell(row_i, j).font = Font(size=9)
                ws8.cell(row_i, j).border = Border(bottom=thin)
            row_i += 1
        for comp in entry.get("expense_components", []):
            ws8.cell(row_i, 1).value = yr
            ws8.cell(row_i, 2).value = mo_name
            ws8.cell(row_i, 3).value = CATEGORY_LABELS.get(comp.get("category", ""), comp.get("category", ""))
            ws8.cell(row_i, 4).value = comp.get("name", "")
            ws8.cell(row_i, 5).value = comp.get("monthly_amount", 0)
            ws8.cell(row_i, 5).number_format = MONEY
            ws8.cell(row_i, 6).value = comp.get("formula", "")
            ws8.cell(row_i, 7).value = "Expense Schedule"
            for j in range(1, 8):
                ws8.cell(row_i, j).font = Font(size=9)
                ws8.cell(row_i, j).border = Border(bottom=thin)
            row_i += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Symbio_Financial_Model_v2.xlsx"},
    )


# ── Plaid helpers ────────────────────────────────────────────────────────────

def _plaid_base() -> str:
    env = os.environ.get("PLAID_ENV", "sandbox")
    return "https://production.plaid.com" if env == "production" else "https://sandbox.plaid.com"


def _plaid_creds() -> dict:
    env = os.environ.get("PLAID_ENV", "sandbox")
    secret = os.environ["PLAID_PRODUCTION_SECRET"] if env == "production" else os.environ["PLAID_SANDBOX_SECRET"]
    return {"client_id": os.environ["PLAID_CLIENT_ID"], "secret": secret}


def _do_plaid_sync() -> dict:
    """Pull balance + last-30-day transactions from Plaid. Stores a row in fpa_actuals."""
    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute(
            "SELECT access_token, excluded_account_ids FROM fpa_plaid_tokens WHERE is_active = true ORDER BY created_at DESC LIMIT 1"
        )
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=400, detail="No Plaid account connected")
        access_token = row["access_token"]
        excluded = set(row["excluded_account_ids"] or [])

        creds = _plaid_creds()
        base = _plaid_base()

        # Current balance (cached — /accounts/balance/get requires the balance product)
        bal = httpx.post(
            f"{base}/accounts/get",
            json={**creds, "access_token": access_token},
            timeout=30,
        )
        if bal.status_code != 200:
            raise HTTPException(status_code=502, detail=f"Plaid accounts error: {bal.text}")
        accounts = bal.json()["accounts"]
        cash_balance = sum(
            float(a["balances"]["current"] or 0)
            for a in accounts
            if a.get("type") == "depository" and a["account_id"] not in excluded
        )

        # Last 30 days of transactions (filtered to included accounts)
        today = date.today()
        start = today - timedelta(days=30)
        txn = httpx.post(
            f"{base}/transactions/get",
            json={
                **creds,
                "access_token": access_token,
                "start_date": start.isoformat(),
                "end_date": today.isoformat(),
            },
            timeout=30,
        )
        if txn.status_code != 200:
            raise HTTPException(status_code=502, detail=f"Plaid transactions error: {txn.text}")
        transactions = [
            t for t in txn.json().get("transactions", [])
            if t["account_id"] not in excluded
        ]

        # Plaid sign convention: positive = money out, negative = money in
        outflow = sum(float(t["amount"]) for t in transactions if float(t["amount"]) > 0)
        inflow = sum(abs(float(t["amount"])) for t in transactions if float(t["amount"]) < 0)
        net_burn = outflow - inflow

        cur.execute(
            """INSERT INTO fpa_actuals (cash_balance, monthly_inflow, monthly_outflow, net_burn, source)
               VALUES (%s, %s, %s, %s, 'plaid')
               RETURNING id, pulled_at""",
            (cash_balance, inflow, outflow, net_burn),
        )
        conn.commit()
        saved = cur.fetchone()
        cur.close()

        return {
            "id": saved["id"],
            "pulled_at": saved["pulled_at"].isoformat(),
            "cash_balance": cash_balance,
            "monthly_inflow": inflow,
            "monthly_outflow": outflow,
            "net_burn": net_burn,
        }
    finally:
        conn.close()


def _compute_kpis(actuals: dict) -> dict:
    capital_adj = float(actuals.get("capital_adjustment", 0) or 0)
    net_burn = float(actuals["net_burn"]) + capital_adj  # add back one-time capital injections
    cash = float(actuals["cash_balance"])
    inflow = float(actuals["monthly_inflow"]) - capital_adj

    runway_months = (cash / net_burn) if net_burn > 0 else None
    zero_date = None
    if runway_months is not None:
        zero_date = (datetime.now() + timedelta(days=runway_months * 30)).date().isoformat()

    return {
        "burn_rate_monthly": round(net_burn, 2),
        "burn_rate_daily": round(net_burn / 30, 2),
        "burn_rate_hourly": round(net_burn / 720, 2),
        "run_rate_monthly": round(inflow, 2),
        "run_rate_daily": round(inflow / 30, 2),
        "run_rate_hourly": round(inflow / 720, 2),
        "runway_months": round(runway_months, 1) if runway_months is not None else None,
        "zero_date": zero_date,
        "cash_balance": round(cash, 2),
    }


# ── Plaid endpoints ──────────────────────────────────────────────────────────

@router.post("/plaid/link-token")
def create_link_token(request: Request):
    """Create a Plaid Link token to initiate bank connection. Admin only."""
    require_admin(request)

    r = httpx.post(
        f"{_plaid_base()}/link/token/create",
        json={
            **_plaid_creds(),
            "client_name": "Collective ERP",
            "country_codes": ["US"],
            "language": "en",
            "user": {"client_user_id": "admin"},
            "products": ["transactions"],
        },
        timeout=30,
    )
    if r.status_code != 200:
        raise HTTPException(status_code=502, detail=f"Plaid error: {r.text}")
    return {"link_token": r.json()["link_token"]}


class PlaidExchangeBody(BaseModel):
    public_token: str
    institution_name: str | None = None


@router.post("/plaid/exchange")
def exchange_plaid_token(body: PlaidExchangeBody, request: Request):
    """Exchange Plaid public token for access token and store it. Admin only."""
    require_admin(request)

    r = httpx.post(
        f"{_plaid_base()}/item/public_token/exchange",
        json={**_plaid_creds(), "public_token": body.public_token},
        timeout=30,
    )
    if r.status_code != 200:
        raise HTTPException(status_code=502, detail=f"Plaid error: {r.text}")

    data = r.json()
    access_token = data["access_token"]
    item_id = data["item_id"]

    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute("UPDATE fpa_plaid_tokens SET is_active = false WHERE is_active = true")
        cur.execute(
            "INSERT INTO fpa_plaid_tokens (access_token, item_id, institution_name) VALUES (%s, %s, %s)",
            (access_token, item_id, body.institution_name),
        )
        conn.commit()
        cur.close()
    finally:
        conn.close()

    return {"status": "connected", "item_id": item_id}


@router.get("/plaid/status")
def plaid_status(request: Request):
    """Return whether a Plaid account is connected. Admin only."""
    require_admin(request)

    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute(
            "SELECT institution_name, created_at FROM fpa_plaid_tokens WHERE is_active = true ORDER BY created_at DESC LIMIT 1"
        )
        row = cur.fetchone()
        cur.close()
    finally:
        conn.close()

    if not row:
        return {"connected": False}
    return {
        "connected": True,
        "institution_name": row["institution_name"],
        "connected_at": row["created_at"].isoformat() if row["created_at"] else None,
    }


@router.get("/plaid/accounts")
def list_plaid_accounts(request: Request):
    """Return connected accounts with balances and exclusion status. Admin only."""
    require_admin(request)

    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute(
            "SELECT access_token, excluded_account_ids FROM fpa_plaid_tokens WHERE is_active = true ORDER BY created_at DESC LIMIT 1"
        )
        row = cur.fetchone()
        cur.close()
    finally:
        conn.close()

    if not row:
        raise HTTPException(status_code=404, detail="No Plaid account connected")

    excluded = set(row["excluded_account_ids"] or [])
    creds = _plaid_creds()
    r = httpx.post(
        f"{_plaid_base()}/accounts/get",
        json={**creds, "access_token": row["access_token"]},
        timeout=30,
    )
    if r.status_code != 200:
        raise HTTPException(status_code=502, detail=f"Plaid error: {r.text}")

    return [
        {
            "account_id": a["account_id"],
            "name": a["name"],
            "type": a["type"],
            "subtype": a["subtype"],
            "balance": float(a["balances"]["current"] or 0),
            "excluded": a["account_id"] in excluded,
        }
        for a in r.json()["accounts"]
    ]


class AccountExclusionUpdate(BaseModel):
    excluded_account_ids: list[str]


@router.patch("/plaid/accounts")
def update_account_exclusions(body: AccountExclusionUpdate, request: Request):
    """Update which accounts are excluded from actuals calculations. Admin only."""
    require_admin(request)

    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute(
            "UPDATE fpa_plaid_tokens SET excluded_account_ids = %s WHERE is_active = true",
            (body.excluded_account_ids,),
        )
        conn.commit()
        cur.close()
    finally:
        conn.close()

    return {"excluded_account_ids": body.excluded_account_ids}


# ── Actuals endpoints ────────────────────────────────────────────────────────

@router.get("/actuals")
def get_actuals(request: Request):
    """Return latest actuals and computed KPIs. Admin only."""
    require_admin(request)

    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute("SELECT * FROM fpa_actuals ORDER BY pulled_at DESC LIMIT 1")
        row = cur.fetchone()
        cur.close()
    finally:
        conn.close()

    if not row:
        return {"actuals": None, "kpis": None}

    actuals = dict(row)
    actuals["pulled_at"] = actuals["pulled_at"].isoformat() if actuals["pulled_at"] else None
    for f in ("cash_balance", "monthly_inflow", "monthly_outflow", "net_burn", "capital_adjustment"):
        if f in actuals and actuals[f] is not None:
            actuals[f] = float(actuals[f])

    return {"actuals": actuals, "kpis": _compute_kpis(actuals)}


@router.post("/actuals/sync")
def sync_actuals(request: Request):
    """Trigger an immediate Plaid sync. Admin only."""
    require_admin(request)
    result = _do_plaid_sync()
    return {"actuals": result, "kpis": _compute_kpis(result)}


class CapitalAdjustmentBody(BaseModel):
    amount: float


@router.patch("/actuals/capital-adjustment")
def set_capital_adjustment(body: CapitalAdjustmentBody, request: Request):
    """Set a one-time capital adjustment on the latest actuals row to exclude from burn rate. Admin only."""
    require_admin(request)

    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute(
            """UPDATE fpa_actuals SET capital_adjustment = %s
               WHERE id = (SELECT id FROM fpa_actuals ORDER BY pulled_at DESC LIMIT 1)
               RETURNING *""",
            (body.amount,),
        )
        conn.commit()
        row = cur.fetchone()
        cur.close()
    finally:
        conn.close()

    if not row:
        raise HTTPException(status_code=404, detail="No actuals row found")

    actuals = dict(row)
    actuals["pulled_at"] = actuals["pulled_at"].isoformat() if actuals["pulled_at"] else None
    for f in ("cash_balance", "monthly_inflow", "monthly_outflow", "net_burn", "capital_adjustment"):
        if f in actuals and actuals[f] is not None:
            actuals[f] = float(actuals[f])

    return {"actuals": actuals, "kpis": _compute_kpis(actuals)}


# ── QBO helpers ──────────────────────────────────────────────────────────────

import base64
import secrets
from datetime import timezone
from urllib.parse import quote
from fastapi.responses import RedirectResponse

QBO_API_BASE = "https://quickbooks.api.intuit.com"
QBO_AUTH_BASE = "https://appcenter.intuit.com/connect/oauth2"
QBO_TOKEN_URL = "https://oauth.platform.intuit.com/oauth2/v1/tokens/bearer"


def _qbo_redirect_uri() -> str:
    base = os.environ.get("NEXTAUTH_URL", "https://platform.collectiveerp.io").rstrip("/")
    return f"{base}/api/fpa/qbo/callback"


def _qbo_basic_auth() -> str:
    cid = os.environ.get("QBO_CLIENT_ID", "")
    csec = os.environ.get("QBO_CLIENT_SECRET", "")
    return base64.b64encode(f"{cid}:{csec}".encode()).decode()


def _qbo_get_access_token(conn) -> tuple[str, str]:
    """Return (access_token, realm_id), refreshing if within 5 min of expiry."""
    cur = conn.cursor()
    cur.execute(
        "SELECT access_token, refresh_token, realm_id, expires_at FROM fpa_qbo_tokens WHERE is_active = true ORDER BY created_at DESC LIMIT 1"
    )
    row = cur.fetchone()
    cur.close()
    if not row:
        raise HTTPException(status_code=400, detail="QuickBooks not connected")

    expires_at = row["expires_at"]
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    now = datetime.now(timezone.utc)

    if (expires_at - now).total_seconds() < 300:
        # Refresh
        r = httpx.post(
            QBO_TOKEN_URL,
            headers={
                "Authorization": f"Basic {_qbo_basic_auth()}",
                "Content-Type": "application/x-www-form-urlencoded",
            },
            data={"grant_type": "refresh_token", "refresh_token": row["refresh_token"]},
            timeout=30,
        )
        if r.status_code != 200:
            raise HTTPException(status_code=502, detail=f"QBO token refresh failed: {r.text}")
        tokens = r.json()
        new_access = tokens["access_token"]
        new_refresh = tokens["refresh_token"]
        new_expires = datetime.now(timezone.utc) + timedelta(seconds=tokens["expires_in"])
        cur2 = conn.cursor()
        cur2.execute(
            "UPDATE fpa_qbo_tokens SET access_token=%s, refresh_token=%s, expires_at=%s WHERE is_active=true",
            (new_access, new_refresh, new_expires),
        )
        conn.commit()
        cur2.close()
        return new_access, row["realm_id"]

    return row["access_token"], row["realm_id"]


def _parse_qbo_pl(report: dict, start: date, summarize_by: str) -> list[dict]:
    """
    Parse QBO P&L report into a list of per-period dicts.
    ColData layout: [label, period1, period2, ..., periodN, total]
    We use ColData[1:-1] for the individual periods, dropping the total.
    """
    all_rows = report.get("Rows", {}).get("Row", [])

    def extract_values(group: str) -> list[float]:
        for row in all_rows:
            if row.get("group") == group:
                col_data = row.get("Summary", {}).get("ColData", [])
                # [1:-1] drops the label (index 0) and the total (last element)
                vals = col_data[1:-1]
                return [float(cd.get("value") or 0) for cd in vals]
        return []

    revenues = extract_values("Income")
    cogs      = extract_values("COGS")
    expenses  = extract_values("Expenses")
    other_exp = extract_values("OtherExpenses")
    net_incomes = extract_values("NetIncome")

    # Total expenses = COGS + operating Expenses + OtherExpenses (all may contain labor)
    n = max(len(revenues), len(expenses), len(net_incomes), len(cogs))
    if n == 0:
        return []

    def _sum(*lists: list[float], idx: int) -> float:
        return sum(lst[idx] if idx < len(lst) else 0.0 for lst in lists)

    results = []
    for i in range(n):
        # Compute period start/end
        import calendar
        if summarize_by == "Week":
            p_start = start + timedelta(weeks=i)
            p_end = p_start + timedelta(days=6)
            label = p_start.strftime("%b %d")
        elif summarize_by == "Quarter":
            # Advance by i quarters from start
            m = start.month + i * 3
            y = start.year + (m - 1) // 12
            m = ((m - 1) % 12) + 1
            p_start = date(y, m, 1)
            end_m = m + 2
            end_y = y + (end_m - 1) // 12
            end_m = ((end_m - 1) % 12) + 1
            p_end = date(end_y, end_m, calendar.monthrange(end_y, end_m)[1])
            q_num = (m - 1) // 3 + 1
            label = f"Q{q_num} {y}"
        elif summarize_by == "Year":
            y = start.year + i
            p_start = date(y, 1, 1)
            p_end = date(y, 12, 31)
            label = str(y)
        else:
            # Monthly
            m = start.month + i
            y = start.year + (m - 1) // 12
            m = ((m - 1) % 12) + 1
            p_start = date(y, m, 1)
            p_end = date(y, m, calendar.monthrange(y, m)[1])
            label = p_start.strftime("%b %Y")
        total_exp = _sum(cogs, expenses, other_exp, idx=i)
        results.append({
            "period_label": label,
            "period_start": p_start,
            "period_end": p_end,
            "revenue": revenues[i] if i < len(revenues) else 0.0,
            "cogs": cogs[i] if i < len(cogs) else 0.0,
            "operating_expenses": expenses[i] if i < len(expenses) else 0.0,
            "expenses": total_exp,  # COGS + OpEx + OtherExp
            "net_income": net_incomes[i] if i < len(net_incomes) else 0.0,
        })
    return results


def _do_qbo_sync(period_type: str = "monthly") -> list[dict]:
    """Pull QBO P&L and upsert into fpa_qbo_periods."""
    conn = get_conn()
    try:
        access_token, realm_id = _qbo_get_access_token(conn)

        import calendar
        today = date.today()
        if period_type == "weekly":
            start = today - timedelta(weeks=12)  # 12 weeks of history
            summarize_by = "Week"
        elif period_type == "quarterly":
            # 8 quarters back (2 years)
            cur_q = (today.month - 1) // 3
            q_start_month = cur_q * 3 + 1
            m = q_start_month - 21  # 7 quarters back
            y = today.year
            while m <= 0:
                m += 12
                y -= 1
            start = date(y, m, 1)
            summarize_by = "Quarter"
        elif period_type == "yearly":
            start = date(today.year - 4, 1, 1)  # 5 years of history
            summarize_by = "Year"
        else:  # monthly
            m = today.month - 11  # 12 months of history
            y = today.year
            while m <= 0:
                m += 12
                y -= 1
            start = date(y, m, 1)
            summarize_by = "Month"

        r = httpx.get(
            f"{QBO_API_BASE}/v3/company/{realm_id}/reports/ProfitAndLoss",
            params={
                "start_date": start.isoformat(),
                "end_date": today.isoformat(),
                "summarize_column_by": summarize_by,
                "minorversion": "65",
            },
            headers={
                "Authorization": f"Bearer {access_token}",
                "Accept": "application/json",
            },
            timeout=30,
        )
        if r.status_code != 200:
            raise HTTPException(status_code=502, detail=f"QBO P&L error: {r.text}")

        periods = _parse_qbo_pl(r.json(), start, summarize_by)

        import json as _json
        cur = conn.cursor()
        # Clear old periods of this type and reinsert fresh
        cur.execute("DELETE FROM fpa_qbo_periods WHERE period_type = %s", (period_type,))
        for p in periods:
            cur.execute(
                """INSERT INTO fpa_qbo_periods (period_type, period_start, period_end, revenue, expenses, net_income, raw_json)
                   VALUES (%s, %s, %s, %s, %s, %s, %s)""",
                (period_type, p["period_start"], p["period_end"],
                 p["revenue"], p["expenses"], p["net_income"], _json.dumps(p, default=str)),
            )
        conn.commit()
        cur.close()
        return periods
    finally:
        conn.close()


# ── QBO OAuth endpoints ──────────────────────────────────────────────────────

@router.get("/qbo/auth-url")
def qbo_auth_url(request: Request):
    """Return Intuit OAuth authorization URL. Admin only."""
    require_admin(request)
    cid = os.environ.get("QBO_CLIENT_ID", "")
    if not cid:
        raise HTTPException(status_code=400, detail="QBO_CLIENT_ID not configured")

    state = secrets.token_urlsafe(16)
    redirect_uri = quote(_qbo_redirect_uri(), safe="")
    params = (
        f"?client_id={cid}"
        f"&redirect_uri={redirect_uri}"
        f"&response_type=code"
        f"&scope=com.intuit.quickbooks.accounting"
        f"&state={state}"
    )
    return {"auth_url": QBO_AUTH_BASE + params, "state": state}


@router.get("/qbo/callback")
def qbo_callback(code: str, realmId: str, state: str | None = None):
    """Handle Intuit OAuth callback, exchange code for tokens."""
    cid = os.environ.get("QBO_CLIENT_ID", "")
    if not cid:
        raise HTTPException(status_code=400, detail="QBO not configured")

    r = httpx.post(
        QBO_TOKEN_URL,
        headers={
            "Authorization": f"Basic {_qbo_basic_auth()}",
            "Content-Type": "application/x-www-form-urlencoded",
        },
        data={
            "grant_type": "authorization_code",
            "code": code,
            "redirect_uri": _qbo_redirect_uri(),
        },
        timeout=30,
    )
    if r.status_code != 200:
        raise HTTPException(status_code=502, detail=f"QBO token exchange failed: {r.text}")

    tokens = r.json()
    access_token = tokens["access_token"]
    refresh_token = tokens["refresh_token"]
    expires_at = datetime.now(timezone.utc) + timedelta(seconds=tokens["expires_in"])

    # Fetch company info for realm_name
    realm_name = None
    try:
        ci = httpx.get(
            f"{QBO_API_BASE}/v3/company/{realmId}/companyinfo/{realmId}",
            headers={"Authorization": f"Bearer {access_token}", "Accept": "application/json"},
            params={"minorversion": "65"},
            timeout=15,
        )
        if ci.status_code == 200:
            realm_name = ci.json().get("CompanyInfo", {}).get("CompanyName")
    except Exception:
        pass

    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute("UPDATE fpa_qbo_tokens SET is_active = false WHERE is_active = true")
        cur.execute(
            """INSERT INTO fpa_qbo_tokens (realm_id, realm_name, access_token, refresh_token, expires_at)
               VALUES (%s, %s, %s, %s, %s)""",
            (realmId, realm_name, access_token, refresh_token, expires_at),
        )
        conn.commit()
        cur.close()
    finally:
        conn.close()

    base = os.environ.get("NEXTAUTH_URL", "https://platform.collectiveerp.io").rstrip("/")
    return RedirectResponse(url=f"{base}/settings?qbo=connected")


@router.get("/qbo/status")
def qbo_status(request: Request):
    """Return QBO connection status. Admin only."""
    require_admin(request)

    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute(
            "SELECT realm_name, created_at FROM fpa_qbo_tokens WHERE is_active = true ORDER BY created_at DESC LIMIT 1"
        )
        token_row = cur.fetchone()
        cur.execute(
            "SELECT MAX(pulled_at) as last_synced FROM fpa_qbo_periods"
        )
        sync_row = cur.fetchone()
        cur.close()
    finally:
        conn.close()

    if not token_row:
        return {"connected": False}
    return {
        "connected": True,
        "realm_name": token_row["realm_name"],
        "connected_at": token_row["created_at"].isoformat() if token_row["created_at"] else None,
        "last_synced": sync_row["last_synced"].isoformat() if sync_row and sync_row["last_synced"] else None,
    }


def _extract_qbo_categories(rows: list, prefix: str = "") -> list[dict]:
    """Recursively extract line-item categories from QBO P&L row list."""
    out = []
    for item in rows:
        itype = item.get("type", "")
        if itype == "Data":
            cd = item.get("ColData", [])
            name = cd[0].get("value", "") if cd else ""
            amt = float(cd[1].get("value") or 0) if len(cd) > 1 else 0.0
            full_name = f"{prefix}{name}" if prefix else name
            if full_name and amt > 0:
                out.append({"category": full_name, "amount": round(amt, 2)})
        elif itype == "Section":
            hdr_cd = item.get("Header", {}).get("ColData", [{}])
            sec_name = hdr_cd[0].get("value", "") if hdr_cd else ""
            sub_rows = item.get("Rows", {}).get("Row", [])
            if sub_rows:
                sub_label = f"{sec_name} › " if sec_name else ""
                child_items = _extract_qbo_categories(sub_rows, prefix=f"{prefix}{sub_label}")
                if child_items:
                    out.extend(child_items)
                else:
                    summary_cd = item.get("Summary", {}).get("ColData", [])
                    amt = float(summary_cd[1].get("value") or 0) if len(summary_cd) > 1 else 0.0
                    full_name = f"{prefix}{sec_name}" if sec_name else prefix.rstrip(" › ")
                    if full_name and amt > 0:
                        out.append({"category": full_name, "amount": round(amt, 2)})
            else:
                summary_cd = item.get("Summary", {}).get("ColData", [])
                amt = float(summary_cd[1].get("value") or 0) if len(summary_cd) > 1 else 0.0
                full_name = f"{prefix}{sec_name}" if sec_name else prefix.rstrip(" › ")
                if full_name and amt > 0:
                    out.append({"category": full_name, "amount": round(amt, 2)})
    return out


def _do_qbo_category_sync() -> list[dict]:
    """Pull YTD expense breakdown by top-level category from QBO."""
    conn = get_conn()
    try:
        access_token, realm_id = _qbo_get_access_token(conn)
        today = date.today()
        start = date(today.year, 1, 1)

        r = httpx.get(
            f"{QBO_API_BASE}/v3/company/{realm_id}/reports/ProfitAndLoss",
            params={
                "start_date": start.isoformat(),
                "end_date": today.isoformat(),
                "summarize_column_by": "Total",
                "minorversion": "65",
            },
            headers={"Authorization": f"Bearer {access_token}", "Accept": "application/json"},
            timeout=30,
        )
        if r.status_code != 200:
            raise HTTPException(status_code=502, detail=f"QBO P&L error: {r.text}")

        rows = r.json().get("Rows", {}).get("Row", [])
        categories = []
        for row in rows:
            grp = row.get("group", "")
            # Include COGS (direct labor/materials) AND operating Expenses AND OtherExpenses
            if grp in ("COGS", "Expenses", "OtherExpenses"):
                grp_label = "COGS › " if grp == "COGS" else ("Other › " if grp == "OtherExpenses" else "")
                sub_items = row.get("Rows", {}).get("Row", [])
                categories.extend(_extract_qbo_categories(sub_items, prefix=grp_label))

        import json as _json
        cur = conn.cursor()
        cur.execute("DELETE FROM fpa_qbo_expense_categories")
        cur.execute(
            "INSERT INTO fpa_qbo_expense_categories (period_start, period_end, categories) VALUES (%s, %s, %s)",
            (start, today, _json.dumps(categories)),
        )
        conn.commit()
        cur.close()
        return categories
    finally:
        conn.close()


def _write_qbo_actuals_to_excel():
    """No-op: the uploaded Excel is kept pristine (formula cache must not be cleared).
    QBO actuals live in fpa_qbo_periods; the model budget lives in fpa_model.
    Comparison happens in the UI — no file modification needed."""
    pass


def _do_qbo_transaction_sync(start: date, end: date) -> list[dict]:
    """Pull all individual transactions from QBO TransactionList report for a date range."""
    conn = get_conn()
    try:
        access_token, realm_id = _qbo_get_access_token(conn)

        r = httpx.get(
            f"{QBO_API_BASE}/v3/company/{realm_id}/reports/TransactionList",
            params={
                "start_date": start.isoformat(),
                "end_date": end.isoformat(),
                "minorversion": "65",
            },
            headers={"Authorization": f"Bearer {access_token}", "Accept": "application/json"},
            timeout=60,
        )
        if r.status_code != 200:
            raise HTTPException(status_code=502, detail=f"QBO TransactionList error: {r.text}")

        report = r.json()

        # Parse column positions from header
        cols = {}
        for col_section in report.get("Columns", {}).get("Column", []):
            meta = col_section.get("MetaData", [])
            for m in meta:
                if m.get("Name") == "ColKey":
                    idx = len(cols)
                    cols[m["Value"]] = idx

        # ColKey names: tx_date, txn_type, doc_num, name, account_name,
        #               other_account (split/category), memo, subt_nat_amount
        def col_idx(key):
            return cols.get(key)

        rows_data = report.get("Rows", {}).get("Row", [])
        transactions = []
        for row in rows_data:
            if row.get("type") != "Data":
                continue
            cd = row.get("ColData", [])
            def val(key):
                idx = col_idx(key)
                return cd[idx].get("value", "") if idx is not None and idx < len(cd) else ""

            raw_date = val("tx_date")
            try:
                txn_date = date.fromisoformat(raw_date)
            except Exception:
                continue

            raw_amt = val("subt_nat_amount")
            try:
                amount = float(raw_amt) if raw_amt else 0.0
            except Exception:
                amount = 0.0

            if amount == 0.0:
                continue

            txn_type = val("txn_type")
            account = val("account_name")   # bank/CC account
            other_account = val("other_account")  # expense category (split)
            name = val("name")
            memo = val("memo")
            doc_num = val("doc_num")

            # Exclude inter-account transfers — not real expenses
            EXCLUDED_TYPES = {
                "Credit Card Payment", "Transfer", "Journal Entry",
                "Deposit",  # Deposits are income, not expenses
            }
            if txn_type in EXCLUDED_TYPES:
                continue

            # Use the split/category account as primary; fall back to bank account
            category = other_account if other_account else account

            # Only track outflow/expense transaction types
            EXPENSE_TYPES = {
                "Bill", "Bill Payment (Check)", "Bill Payment (Credit Card)",
                "Check", "Credit Card Charge", "Expense", "Purchase",
                "Vendor Credit", "Credit Card Credit",
            }
            if txn_type not in EXPENSE_TYPES:
                continue

            is_expense = True

            # Normalize: positive = money out, negative = refund/credit
            # QBO: checking outflows are negative, CC charges are positive
            if txn_type == "Credit Card Credit":
                # CC Credits where the category is a bank/CC account are
                # misclassified payments/transfers — exclude them
                if category in ("Checking", "Savings", "J. NOTRICA (5309) - 2") or \
                        category.startswith("Checking") or category.startswith("Credit Card"):
                    continue
                # Legitimate refund: reduces expenses (negative)
                amount = -abs(amount)
            else:
                # All other expense types: money leaving the company (positive)
                amount = abs(amount)

            transactions.append({
                "txn_date": txn_date.isoformat(),
                "txn_type": txn_type,
                "txn_id": doc_num,
                "account": account,
                "category": category,
                "name": name,
                "memo": memo,
                "amount": round(amount, 2),
                "is_expense": is_expense,
            })

        # Upsert into DB — delete range and re-insert
        cur = conn.cursor()
        cur.execute(
            "DELETE FROM fpa_qbo_transactions WHERE txn_date >= %s AND txn_date <= %s",
            (start, end),
        )
        if transactions:
            psycopg2.extras.execute_values(
                cur,
                """INSERT INTO fpa_qbo_transactions
                   (txn_date, txn_type, txn_id, account, category, name, memo, amount, is_expense)
                   VALUES %s""",
                [
                    (t["txn_date"], t["txn_type"], t["txn_id"], t["account"],
                     t["category"], t["name"], t["memo"], t["amount"], t["is_expense"])
                    for t in transactions
                ],
            )
        conn.commit()
        cur.close()
        return transactions
    finally:
        conn.close()


@router.post("/qbo/transactions/sync")
def sync_qbo_transactions(request: Request, start_date: str, end_date: str):
    """Sync individual QBO transactions for a date range. Admin only."""
    require_admin(request)
    try:
        start = date.fromisoformat(start_date)
        end = date.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD.")
    txns = _do_qbo_transaction_sync(start, end)
    return {"synced": len(txns), "start": start.isoformat(), "end": end.isoformat()}


@router.get("/qbo/transactions")
def get_qbo_transactions(
    request: Request,
    start_date: str | None = None,
    end_date: str | None = None,
    category: str | None = None,
):
    """Return stored QBO transactions with optional filters. Admin only."""
    require_admin(request)
    conn = get_conn()
    try:
        cur = conn.cursor()
        conditions = []
        params: list = []
        if start_date:
            conditions.append("txn_date >= %s")
            params.append(start_date)
        if end_date:
            conditions.append("txn_date <= %s")
            params.append(end_date)
        if category:
            conditions.append("category ILIKE %s")
            params.append(f"%{category}%")
        where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
        cur.execute(
            f"""SELECT txn_date, txn_type, txn_id, account, category, name, memo, amount, is_expense
                FROM fpa_qbo_transactions {where}
                ORDER BY txn_date DESC""",
            params,
        )
        rows = cur.fetchall()
        cur.close()
    finally:
        conn.close()
    return {"transactions": [dict(r) for r in rows]}


@router.get("/qbo/transactions/summary")
def get_qbo_transactions_summary(
    request: Request,
    start_date: str | None = None,
    end_date: str | None = None,
    group_by: str = "category",  # "category" | "account" | "month"
):
    """Return QBO transactions aggregated by category/account/month. Admin only."""
    require_admin(request)
    conn = get_conn()
    try:
        cur = conn.cursor()
        conditions = ["amount != 0"]
        params: list = []
        if start_date:
            conditions.append("txn_date >= %s")
            params.append(start_date)
        if end_date:
            conditions.append("txn_date <= %s")
            params.append(end_date)
        where = "WHERE " + " AND ".join(conditions)

        if group_by == "month":
            cur.execute(
                f"""SELECT TO_CHAR(txn_date, 'YYYY-MM') as period,
                           SUM(amount) as total,
                           COUNT(*) as txn_count
                    FROM fpa_qbo_transactions {where}
                    GROUP BY period ORDER BY period""",
                params,
            )
        elif group_by == "account":
            cur.execute(
                f"""SELECT account as label,
                           SUM(amount) as total,
                           COUNT(*) as txn_count
                    FROM fpa_qbo_transactions {where}
                    GROUP BY account ORDER BY total DESC""",
                params,
            )
        else:  # category
            cur.execute(
                f"""SELECT category as label,
                           SUM(amount) as total,
                           COUNT(*) as txn_count
                    FROM fpa_qbo_transactions {where}
                    GROUP BY category ORDER BY total DESC""",
                params,
            )
        rows = cur.fetchall()
        cur.close()
    finally:
        conn.close()
    return {"summary": [dict(r) for r in rows], "group_by": group_by}


@router.post("/qbo/sync")
def sync_qbo(request: Request):
    """Trigger immediate QBO P&L + transaction sync. Admin only."""
    require_admin(request)
    monthly = _do_qbo_sync("monthly")
    weekly = _do_qbo_sync("weekly")
    quarterly = _do_qbo_sync("quarterly")
    yearly = _do_qbo_sync("yearly")
    _do_qbo_category_sync()
    _write_qbo_actuals_to_excel()
    # Also sync full transaction history (2 years back through today)
    txn_start = date(date.today().year - 2, 1, 1)
    try:
        _do_qbo_transaction_sync(txn_start, date.today())
    except Exception as e:
        logger.warning(f"Transaction sync failed (non-fatal): {e}")
    return {"monthly": monthly, "weekly": weekly, "quarterly": quarterly, "yearly": yearly}


@router.get("/qbo/categories")
def get_qbo_categories(request: Request):
    """Return latest YTD expense breakdown by category. Admin only."""
    require_admin(request)
    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute("SELECT categories, period_start, period_end, pulled_at FROM fpa_qbo_expense_categories ORDER BY pulled_at DESC LIMIT 1")
        row = cur.fetchone()
        cur.close()
    finally:
        conn.close()
    if not row:
        return {"categories": [], "period_start": None, "period_end": None}
    return {
        "categories": row["categories"],
        "period_start": row["period_start"].isoformat(),
        "period_end": row["period_end"].isoformat(),
        "pulled_at": row["pulled_at"].isoformat(),
    }


@router.get("/qbo/categories/query")
def query_qbo_categories(
    request: Request,
    start_date: str,
    end_date: str,
):
    """Fetch expense categories from QBO for an arbitrary date range. Admin only."""
    require_admin(request)
    try:
        start = date.fromisoformat(start_date)
        end = date.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD.")
    if end < start:
        raise HTTPException(status_code=400, detail="end_date must be >= start_date")

    conn = get_conn()
    try:
        access_token, realm_id = _qbo_get_access_token(conn)
    finally:
        conn.close()

    r = httpx.get(
        f"{QBO_API_BASE}/v3/company/{realm_id}/reports/ProfitAndLoss",
        params={
            "start_date": start.isoformat(),
            "end_date": end.isoformat(),
            "summarize_column_by": "Total",
            "minorversion": "65",
        },
        headers={"Authorization": f"Bearer {access_token}", "Accept": "application/json"},
        timeout=30,
    )
    if r.status_code != 200:
        raise HTTPException(status_code=502, detail=f"QBO error: {r.text}")

    all_rows = r.json().get("Rows", {}).get("Row", [])
    categories = []
    for row in all_rows:
        grp = row.get("group", "")
        if grp in ("COGS", "Expenses", "OtherExpenses"):
            grp_label = "COGS › " if grp == "COGS" else ("Other › " if grp == "OtherExpenses" else "")
            sub_items = row.get("Rows", {}).get("Row", [])
            categories.extend(_extract_qbo_categories(sub_items, prefix=grp_label))

    return {
        "categories": categories,
        "period_start": start.isoformat(),
        "period_end": end.isoformat(),
    }


@router.get("/qbo/pl-raw")
def qbo_pl_raw(request: Request, start_date: str, end_date: str):
    """Return raw QBO P&L row groups for debugging. Admin only."""
    require_admin(request)
    try:
        start = date.fromisoformat(start_date)
        end = date.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")
    conn = get_conn()
    try:
        access_token, realm_id = _qbo_get_access_token(conn)
    finally:
        conn.close()
    r = httpx.get(
        f"{QBO_API_BASE}/v3/company/{realm_id}/reports/ProfitAndLoss",
        params={"start_date": start.isoformat(), "end_date": end.isoformat(),
                "summarize_column_by": "Total", "minorversion": "65"},
        headers={"Authorization": f"Bearer {access_token}", "Accept": "application/json"},
        timeout=30,
    )
    if r.status_code != 200:
        raise HTTPException(status_code=502, detail=f"QBO error: {r.text}")
    rows = r.json().get("Rows", {}).get("Row", [])
    # Return group names and their top-level summary values
    summary = []
    for row in rows:
        grp = row.get("group", row.get("type", "?"))
        sum_cd = row.get("Summary", {}).get("ColData", [])
        label = sum_cd[0].get("value", grp) if sum_cd else grp
        amt = float(sum_cd[1].get("value") or 0) if len(sum_cd) > 1 else None
        sub_count = len(row.get("Rows", {}).get("Row", []))
        summary.append({"group": grp, "label": label, "total": amt, "sub_rows": sub_count})
    return {"groups": summary, "full": r.json()}


@router.get("/qbo/actuals")
def get_qbo_actuals(request: Request, period_type: str = "monthly"):
    """Return latest QBO period actuals. Admin only."""
    require_admin(request)

    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute(
            """SELECT p.period_type, p.raw_json, p.pulled_at
               FROM fpa_qbo_periods p
               WHERE p.id IN (
                   SELECT MAX(id) FROM fpa_qbo_periods
                   WHERE period_type = %s
                   GROUP BY period_start
               )
               ORDER BY p.period_start""",
            (period_type,),
        )
        rows = cur.fetchall()
        cur.close()
    finally:
        conn.close()

    periods = []
    for row in rows:
        p = dict(row["raw_json"]) if row["raw_json"] else {}
        p["pulled_at"] = row["pulled_at"].isoformat() if row["pulled_at"] else None
        periods.append(p)
    return periods


# ── Disconnect endpoints ─────────────────────────────────────────────────────

@router.delete("/plaid/disconnect")
def disconnect_plaid(request: Request):
    """Deactivate Plaid connection. Admin only."""
    require_admin(request)
    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute("UPDATE fpa_plaid_tokens SET is_active = false WHERE is_active = true")
        conn.commit()
        cur.close()
    finally:
        conn.close()
    return {"status": "disconnected"}


@router.delete("/qbo/disconnect")
def disconnect_qbo(request: Request):
    """Deactivate QBO connection. Admin only."""
    require_admin(request)
    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute("UPDATE fpa_qbo_tokens SET is_active = false WHERE is_active = true")
        conn.commit()
        cur.close()
    finally:
        conn.close()
    return {"status": "disconnected"}
